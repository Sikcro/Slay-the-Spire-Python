from tkinter import *
from CharacterSelection import selectionscreen
from Deck import Deckstart
import pygame
from openpyxl import workbook, load_workbook
sn=Tk()
sn.geometry("680x560")
sn.configure(bg="Dark grey")
def start():
    startgame.pack_forget()
    deckb.pack_forget()
    Leaderboards.pack_forget()
    sn.destroy()
    selectionscreen()



def update_character_anim(ind):
    global Lframes
    global LframeCnt
    global mainguy
    global firstbscreen

    frame = Lframes[ind]
    ind += 1

    if ind == LframeCnt:
        ind = 0
    fire.configure(image=frame)
    sn.after(90, update_character_anim, ind)



pygame.mixer.init()
def music():
    pygame.mixer.music.load("Audio/ken hum.mp3")
    pygame.mixer.music.play(loops=20)

def leaderboard():
    print("hi")

def deck():
    Deckstart()

def Cstarter():
    global pbL
    global psL
    theCoin = psL['B1']
    theCoin.value=0
    psL['B4'].value = "False"
    pbL.save('Excel files/Player Data.xlsx')

def Dstarter():
    global wb2

    ws2['A2'].value = "Strike"
    ws2['A3'].value = "Strike"
    ws2['A3'].value = "Strike"
    ws2['A4'].value = "Strike"
    ws2['A5'].value = "Strike"
    ws2['A6'].value = "Block"
    ws2['A7'].value = "Block"
    ws2['A8'].value = "Block"
    ws2['A9'].value = "Block"
    ws2['A10'].value = "."
    ws2['A11'].value = "."
    ws2['A12'].value = "."
    ws2['A13'].value = "."
    ws2['A14'].value = "."
    ws2['A15'].value = "."

    ws2['B10'].value = "."
    ws2['B11'].value = "."
    ws2['B12'].value = "."
    ws2['B13'].value = "."
    ws2['B14'].value = "."
    ws2['B15'].value = "."
    ws2['E2'].value = "normal"
    wb2.save('Excel files/Active Deck.xlsx')
    Cstarter()


LframeCnt = 12
Lframes = [PhotoImage(file='images/fire.gif', format='gif -index %i' % (i)) for i in range(LframeCnt)]
startbuttonpic= PhotoImage(file="images/start-button.png")
logo = PhotoImage(file="images/slay logo.png")

logolabel = Label(image=logo)
logolabel.place(x=0,y=0)
deckb=Button(sn,text="Deck",command=deck,font=("Times New Roman", 20, "italic"),compound="center",image=startbuttonpic)
deckb.place(relx=0,y=215)
startgame=Button(sn,text="Start",font=("Times New Roman", 20, "italic"),compound="center",command=start,fg="black",image=startbuttonpic)
startgame.place(relx=0,y=330)
Leaderboards=Button(sn,text="Leader\nboard",command=leaderboard,font=("Times New Roman", 20, "italic"),compound="center",image=startbuttonpic)
Leaderboards.place(relx=0,y=440)
fire=Label()
fire.place(relx=0.55,y=230)
sn.after(0, update_character_anim, 0)
wb2 = load_workbook("Excel files/Active Deck.xlsx")
ws2 = wb2.active
pbL = load_workbook("Excel files/Player Data.xlsx")
psL = pbL.active
Dstarter()




music()

sn.mainloop()