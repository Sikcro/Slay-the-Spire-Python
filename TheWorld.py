from tkinter import *
from Deck import Deckstart
from firstBattle import firstbattlestart
from openpyxl import workbook, load_workbook
from secondbattle import secondbattlestart
from QuestionMarkRoom import startQroom
from shop import entershop
import tkinter.messagebox
import pygame
import random
from tkinter.tix import *

def endoftheworld():
    global ns
    ns.destroy()

def Deck():
    Deckstart()



def cardsize():
    global cszcounter
    global cardisnormal
    global wbl
    global wsL
    global cardsizeconfirm


    if cardisnormal ==True:

        if cszcounter ==1:
            thecardsizecell = wsL['A2']
            thecardsizecell.value = "Reskin"
            cardsizeconfirm.config(text="Revert Cards")
            wbL.save('Excel files/Settings.xlsx')

            cszcounter=2

        else:
            thecardsizecell = wsL['A2']
            thecardsizecell.value = "normal"
            wbL.save('Excel files/Settings.xlsx')
            cardsizeconfirm.config(text="Change Cards")
            cszcounter=1







def previouspage():
    global my_spin
    global my_spin2
    global stopmusic
    global setscreen
    global bengalimusicchange
    global bengalilanglabel
    global bengalregionchange
    global japanesemusicchange
    global japaneselanglabel
    global japaneseregion
    global Musiclabel
    global Languagelabel
    global RegionLabel
    global languagebuttonconfirm
    global Regionbuttonconfirm
    global settingslabel

    settingspage1= PhotoImage(file="images/Settings page 1.png")

    try:
        stopmusic.place(relx=0.40, y=150)
        my_spin.place(relx=0.5, y=215)
        my_spin2.place(relx=0.33, y=300)
        languagebuttonconfirm.place(relx=0.85, y=216)
        Regionbuttonconfirm.place(relx=0.68, y=299)
    except:
        print("Not working")

    try:
        japaneseregion.place(relx=0.1, y=267)
        japaneselanglabel.place(relx=0.24, y=190)
        japanesemusicchange.place(relx=0.129, y=125)
    except:
        print("jap not on")

    try:
        bengalregionchange.place(relx=0.0,y=267)
        bengalilanglabel.place(relx=0.24,y=190)
        bengalimusicchange.place(relx=0.129,y=125)

    except:
        print("beng not on")

    try:
        Languagelabel.place(relx=0.129, y=200)
        RegionLabel.place(relx=0.08, y=270)
        Musiclabel.place(relx=0.15, y=130)

    except:
        print("error did not delete")

    settingslabel.config(image=settingspage1)





    setscreen.mainloop()

def nextpage():
    global my_spin
    global my_spin2
    global stopmusic
    global setscreen
    global bengalimusicchange
    global bengalilanglabel
    global bengalregionchange
    global japanesemusicchange
    global japaneselanglabel
    global japaneseregion
    global Musiclabel
    global Languagelabel
    global RegionLabel
    global languagebuttonconfirm
    global Regionbuttonconfirm
    global settingslabel
    global cardsizeconfirm
    mycolor = '#%02x%02x%02x' % (236, 228, 180)
    settingspage2= PhotoImage(file="images/Settings page 2.png")

    try:
        Musiclabel.place_forget()
        Languagelabel.place_forget()
        RegionLabel.place_forget()
    except:
        print("eng not on")

    try:
        japanesemusicchange.place_forget()
        japaneselanglabel.place_forget()
        japaneseregion.place_forget()
    except:
        print("jap not on")

    try:
        bengalimusicchange.place_forget()
        bengalilanglabel.place_forget()
        bengalregionchange.place_forget()

    except:
        print("beng not on")

    try:
        my_spin.place_forget()
        my_spin2.place_forget()
        stopmusic.place_forget()
        languagebuttonconfirm.place_forget()
        Regionbuttonconfirm.place_forget()

    except:
        print("error did not delete")

    settingslabel.config(image=settingspage2)

    EnlargeCards = Label(setscreen, text="Card Art:", fg="grey", bg=mycolor, font=("Calibri", 30))
    EnlargeCards.place(relx=0.15, y=130)
    cardsizeconfirm= Button(setscreen,text="Change Cards",command=cardsize)
    cardsizeconfirm.place(relx=0.5,y=150)



    setscreen.mainloop()




def Regionchange():
    global my_spin2
    global setscreen


    spinboxvar2=my_spin2.get()

    if spinboxvar2 =="Europe":
        randomping=random.randint(10,25)
        tkinter.messagebox.showinfo("Region Changed", "Europe:Ping"+str(randomping))

    if spinboxvar2 =="Asia":
        randomping=random.randint(30,60)
        tkinter.messagebox.showinfo("Region Changed", "Asia:Ping"+str(randomping))

    if spinboxvar2 =="Canada":
        randomping=random.randint(25,40)
        tkinter.messagebox.showinfo("Region Changed", "Canada:Ping"+str(randomping))

    else:
        print("Error region not changed")
def languagechange():
    global setscreen
    global sinboxvar
    global spinboxvar
    global my_spin
    global bengalimusicchange
    global bengalilanglabel
    global bengalregionchange
    global japanesemusicchange
    global japaneselanglabel
    global japaneseregion
    global Musiclabel
    global Languagelabel
    global RegionLabel
    global button
    global button2
    global lvl1
    global wsL
    global wbL
    mycolor = '#%02x%02x%02x' % (236, 228, 180)
    spinboxvar = my_spin.get()
    print(spinboxvar)

    if spinboxvar =="Bengali":
        try:
            Musiclabel.place_forget()
            Languagelabel.place_forget()
            RegionLabel.place_forget()
        except:
            print("Does not exist")

        try:
            japanesemusicchange.place_forget()
            japaneselanglabel.place_forget()
            japaneseregion.place_forget()
        except:
            print("does not exist")

        bengalimusicchange=Label(setscreen,text="সঙ্গীত:",font=("Calibri",36),fg="Grey",bg=mycolor)
        bengalimusicchange.place(relx=0.129,y=125)

        bengalilanglabel=Label(setscreen,text="ভাষা:",font=("Calibri",39),fg="Grey",bg=mycolor)
        bengalilanglabel.place(relx=0.24,y=190)

        bengalregionchange=Label(setscreen,text="অঞ্চল:",font=("Calibri",39),fg="Grey",bg=mycolor)
        bengalregionchange.place(relx=0.0,y=267)

        thelanguagecell=wsL['A1']
        thelanguagecell.value="Bengali"
        wbL.save('Excel files/Settings.xlsx')

        button.config(text="ডেক")
        button2.config(text="সেটিংস")
        lvl1.config(text="স্তর 1")


    if spinboxvar=="Japanese":

        try:
            Musiclabel.place_forget()
            Languagelabel.place_forget()
            RegionLabel.place_forget()
        except:
            print("Does not exist")

        try:
            bengalimusicchange.place_forget()
            bengalilanglabel.place_forget()
            bengalregionchange.place_forget()

        except:
            print("does not exist")
        japanesemusicchange=Label(setscreen,text="音楽:",font=("Calibri",40),fg="Grey",bg=mycolor)
        japanesemusicchange.place(relx=0.129,y=125)

        japaneselanglabel=Label(setscreen,text="言語:",font=("Calibri",39),fg="Grey",bg=mycolor)
        japaneselanglabel.place(relx=0.24,y=190)

        japaneseregion=Label(setscreen,text="領域:",font=("Calibri",39),fg="Grey",bg=mycolor)
        japaneseregion.place(relx=0.1,y=267)

        thelanguagecell=wsL['A1']
        thelanguagecell.value="Japanese"
        wbL.save('Excel files/Settings.xlsx')

        button.config(text="デッキ")
        button2.config(text="設定")
        lvl1.config(text="レベル1")

    if spinboxvar=="English":
        try:
            bengalimusicchange.place_forget()
            bengalilanglabel.place_forget()
            bengalregionchange.place_forget()
        except:
            print("does not exist")
        try:
            japanesemusicchange.place_forget()
            japaneselanglabel.place_forget()
            japaneseregion.place_forget()
        except:
            print("does not exist")

        Musiclabel = Label(setscreen, text="Music:", fg="grey", bg=mycolor, font=("Calibri", 30))
        Musiclabel.place(relx=0.15, y=130)
        Languagelabel = Label(setscreen, text="Language:", fg="grey", bg=mycolor, font=("Calibri", 30))
        Languagelabel.place(relx=0.129, y=200)
        RegionLabel = Label(setscreen, text="Region:", fg="grey", bg=mycolor, font=("Calibri", 30))
        RegionLabel.place(relx=0.08, y=270)

        thelanguagecell=wsL['A1']
        thelanguagecell.value="English"
        wbL.save('Excel files/Settings.xlsx')


        button.config(text="Deck")
        button2.config(text="Settings")
        lvl1.config(text="Lvl1")

    setscreen.mainloop()

def coinslvl3clear():
    global pbL
    global psL
    coincurrentval=psL['B2'].value
    coins1.config(text=coincurrentval)

def coinslvl2clear():
    coins1.config(text="195")

def coinslvl1clear():
    coins1.config(text="075")
def coins():
    global ns
    global pbL
    global psL
    global coins1
    print("enter coins")
    coincolourbg = '#%02x%02x%02x' % (61,73,79)
    coinbag=PhotoImage(file="images/coinbag.png")
    coinbaglabel=Label(ns,image=coinbag)
    coinbaglabel.place(relx=0,y=0)
    if psL['B1'].value ==0:
        coins1 = Label(ns, text="000", font=("Times New Roman", 30,), bg=coincolourbg, fg="yellow")
        coins1.place(relx=0.1, y=2)

    ns.mainloop()


def check():
    global lvl1clicked
    global lvl2clicked
    global lvl3clicked
    global lvl4clicked
    global lvl5clicked

    if lvl1clicked ==0:
        print("success")
        lvl1clicked=1
        firstbattlestart()

    if lvl2clicked ==0:
        print("success")
        lvl2clicked=1
        secondbattlestart()

    if lvl3clicked ==0:
        print("success")
        lvl3clicked=1
        entershop()

    if lvl4clicked ==0:
        print("success")
        startQroom()
        lvl4clicked=1

    if lvl5clicked ==0:
        print("success")
        startQroom()
        lvl5clicked=1


def THEboard():
    global ns
    global mainarea
    global lvl1clicked
    global lvl2clicked
    global lvl3clicked
    global lvl4clicked
    global lvl5clicked
    global lvl1
    global lvl2
    global lvl3
    global lvl4
    global lvl5
    global levelcounter
    global firstrun
    global pbL
    global psL


    levelcounter=levelcounter+1

    if levelcounter==1:
        lvl1=Button(mainarea,text="stage1",command=check)
        lvl1.place(relx=0.43, y=810, anchor="center")
        lvl1clicked=0


    if levelcounter==2:
        lvl1.place_forget()
        lvl2=Button(mainarea,text="stage2",command=check)
        lvl2.place(relx=0.43, y=710, anchor="center")
        lvl2clicked=0

    if levelcounter==3:
        lvl2.place_forget()
        lvl3=Button(mainarea,text="shop",command=check)
        lvl3.place(relx=0.45, y=600, anchor="center")
        lvl3clicked=0

    if levelcounter==4:
        lvl3.place_forget()
        lvl4=Button(mainarea,text="???",command=check)
        lvl4.place(relx=0.45, y=500, anchor="center")
        lvl4clicked=0

    if levelcounter==5:
        lvl4.place_forget()
        lvl5=Button(mainarea,text="MINI BOSS",command=check)
        lvl5.place(relx=0.45, y=340, anchor="center")
        lvl5clicked=0

    if firstrun==True:
        ns.after(100, coins)
        firstrun=False
    ns.mainloop()


def stopthemusic():
    global musicison
    global stopmusic
    global oncounter


    if musicison ==True:

        if oncounter ==1:
            pygame.mixer.music.stop()
            stopmusic.configure(text="ON")
            oncounter=2

        else:
            pygame.mixer.music.play()
            stopmusic.configure(text="OFF")
            oncounter=1


def settings():
    global musicison
    global stopmusic
    global oncounter
    global spinboxvar
    global my_spin
    global setscreen
    global Musiclabel
    global Languagelabel
    global my_spin2
    global RegionLabel
    global languagebuttonconfirm
    global Regionbuttonconfirm
    global settingslabel
    global cszcounter
    global cardisnormal
    global wsL
    global wbL
    setscreen=Toplevel()
    setscreen.geometry("550x450")
    mycolor = '#%02x%02x%02x' % (236, 228, 180)
    settingsbg = PhotoImage(file="images/Settings page 1.png")
    rightarrow = PhotoImage(file="images/right arrow.png")
    leftarrow = PhotoImage(file="images/left arrow.png")
    musicison = True
    cardisnormal = True
    settingslabel = Label(setscreen,image=settingsbg)
    settingslabel.place(relx=0,y=0)

    Musiclabel=Label(setscreen,text="Music:",fg="grey",bg=mycolor,font=("Calibri",30))
    Musiclabel.place(relx=0.15,y=130)


    rightarrowbutton = Button(setscreen,image=rightarrow,command=nextpage)
    rightarrowbutton.place(relx=0.8,y=29)

    leftarrowbutton = Button(setscreen, image=leftarrow,command=previouspage)
    leftarrowbutton.place(relx=0.02, y=29)

    Languagelabel=Label(setscreen,text="Language:",fg="grey",bg=mycolor,font=("Calibri",30))
    Languagelabel.place(relx=0.129,y=200)
    my_spin = Spinbox(setscreen, width=30,value=("English","Bengali","Japanese"))
    my_spin.place(relx=0.5,y=215)
    languagebuttonconfirm=Button(setscreen,text="Enter",command=languagechange)
    languagebuttonconfirm.place(relx=0.85,y=216)

    RegionLabel=Label(setscreen,text="Region:",fg="grey",bg=mycolor,font=("Calibri",30))
    RegionLabel.place(relx=0.08,y=270)
    my_spin2 = Spinbox(setscreen, width=30,value=("Europe","Asia","Canada"))
    my_spin2.place(relx=0.33,y=300)
    Regionbuttonconfirm=Button(setscreen,text="Enter",command=Regionchange)
    Regionbuttonconfirm.place(relx=0.68,y=299)

    oncounter=1
    cszcounter=1
    stopmusic= Button(setscreen,text="OFF",command=stopthemusic)
    stopmusic.place(relx=0.40,y=150)

    thecardsizecell = wsL['A2']
    thecardsizecell.value = "normal"
    wbL.save('Excel files/Settings.xlsx')




    setscreen.mainloop()


def gamestart():
    global ns
    global mainarea
    global button
    global button2
    global wsL
    global wbL
    global pbL
    global psL
    global levelcounter
    global coinsval
    global firstrun
    ns=Tk()
    levelcounter=0
    firstrun = True
    mainarea = Frame(ns, bg='#CCC', width=600, height=850)
    mainarea.pack(expand=True, fill='both', side='left')
    sidebar = Frame(ns, width=400, bg='Dark grey', height=300, relief='sunken', borderwidth=2)
    sidebar.pack(expand=False, fill='both', side='right', anchor='nw')
    # Add image file
    bg = PhotoImage(file="images/map.png")
    # Show image using label
    label1 = Label(mainarea, image=bg)
    label1.place(x=0, y=0)

    button = Button(sidebar,width=10,height=10, text="Deck",fg="white",bg="Dark Blue",command=Deck)
    button.pack()
    button2 = Button(sidebar,width=10,height=10, text="Settings",fg="white",bg="Dark Grey",command=settings)
    button2.pack()

    wbL = load_workbook("Excel files/Settings.xlsx")
    wsL=wbL.active
    pbL = load_workbook("Excel files/Player Data.xlsx")
    psL=pbL.active

    THEboard()
