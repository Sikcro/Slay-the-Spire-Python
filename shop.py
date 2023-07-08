from tkinter import *
from openpyxl import workbook, load_workbook
from PIL import Image, ImageTk
from itertools import count
import random





def exit():
    global shopsn
    from TheWorld import THEboard
    from TheWorld import coinslvl3clear
    global coinshopvar
    global psL
    psL['B2'].value = coinshopvar
    pbL.save('Excel files/Player Data.xlsx')
    shopsn.after(100,coinslvl3clear)
    shopsn.after(300,shopsn.destroy)
    THEboard()

def cardRemover():
    global wb2
    global ws2
    global coins
    global coinshopvar
    from Deck import Deckstart
    removeprice=120
    if removeprice > coinshopvar:
        print("Insufficient funds")
    else:
        coinshopvar=coinshopvar-removeprice
        ws2['E2'].value = "Delete"
        coins.config(text=coinshopvar)
        wb2.save('Excel files/Active Deck.xlsx')

        Deckstart()

def cardselection1():
    global wb2
    global ws2
    global rscreen
    global price
    global coinshopvar
    global purchacelimit
    print("col 1 working")

    if price > coinshopvar:
        print("Insufficient funds")
    else:
        purchacelimit=purchacelimit+1

        if purchacelimit ==1:
            coinshopvar=coinshopvar-price
            coins.config(text=coinshopvar)
            thercard = ws2['A12']
            thercard.value = "multi"
            wb2.save('Excel files/Active Deck.xlsx')

        if purchacelimit ==2:
            coinshopvar=coinshopvar-price
            coins.config(text=coinshopvar)
            thercard = ws2['A13']
            thercard.value = "multi"
            wb2.save('Excel files/Active Deck.xlsx')

        if purchacelimit ==3:
            coinshopvar=coinshopvar-price
            coins.config(text=coinshopvar)
            thercard = ws2['A14']
            thercard.value = "multi"
            wb2.save('Excel files/Active Deck.xlsx')

        if purchacelimit ==4:
            coinshopvar=coinshopvar-price
            coins.config(text=coinshopvar)
            thercard = ws2['A15']
            thercard.value = "multi"
            wb2.save('Excel files/Active Deck.xlsx')



def cardselection2():
    global wb2
    global ws2
    global rscreen
    global price
    global coinshopvar
    global purchacelimit

    if price > coinshopvar:
        print("Insufficient funds")
    else:
        purchacelimit=purchacelimit+1
        coinshopvar=coinshopvar-price
        coins.config(text=coinshopvar)
        if purchacelimit ==1:
            print("col 2 working")
            thercard = ws2['A12']
            thercard.value = "repair"
            wb2.save('Excel files/Active Deck.xlsx')

        if purchacelimit ==2:
            print("col 2 working")
            thercard = ws2['A13']
            thercard.value = "repair"
            wb2.save('Excel files/Active Deck.xlsx')

        if purchacelimit ==3:
            print("col 2 working")
            thercard = ws2['A14']
            thercard.value = "repair"
            wb2.save('Excel files/Active Deck.xlsx')

        if purchacelimit ==4:
            print("col 2 working")
            thercard = ws2['A15']
            thercard.value = "repair"
            wb2.save('Excel files/Active Deck.xlsx')




def cardselection3():
    global wb2
    global ws2
    global rscreen
    global price
    global coinshopvar
    global purchacelimit

    if price > coinshopvar:
        print("Insufficient funds")
    else:
        coinshopvar=coinshopvar-price
        coins.config(text=coinshopvar)
        purchacelimit=purchacelimit+1

        if purchacelimit ==1:
            print("col 3 working")
            thercard = ws2['A12']
            thercard.value = "heavy"
            wb2.save('Excel files/Active Deck.xlsx')

        if purchacelimit ==2:
            print("col 3 working")
            thercard = ws2['A13']
            thercard.value = "heavy"
            wb2.save('Excel files/Active Deck.xlsx')

        if purchacelimit ==3:
            print("col 3 working")
            thercard = ws2['A14']
            thercard.value = "heavy"
            wb2.save('Excel files/Active Deck.xlsx')

        if purchacelimit ==4:
            print("col 3 working")
            thercard = ws2['A12']
            thercard.value = "heavy"
            wb2.save('Excel files/Active Deck.xlsx')




def cardselection5():
    global wb2
    global ws2
    global rscreen
    global price
    global coinshopvar
    global purchacelimit
    if price > coinshopvar:
        print("Insufficient funds")
    else:
        purchacelimit=purchacelimit+1
        coinshopvar=coinshopvar-price
        coins.config(text=coinshopvar)

        if purchacelimit==1:
            print("col 5 working")
            thercard = ws2['A12']
            thercard.value = "30block"
            wb2.save('Excel files/Active Deck.xlsx')

        if purchacelimit==2:
            print("col 5 working")
            thercard = ws2['A13']
            thercard.value = "30block"
            wb2.save('Excel files/Active Deck.xlsx')

        if purchacelimit==3:
            print("col 5 working")
            thercard = ws2['A14']
            thercard.value = "30block"
            wb2.save('Excel files/Active Deck.xlsx')

        if purchacelimit==4:
            print("col 5 working")
            thercard = ws2['A15']
            thercard.value = "30block"
            wb2.save('Excel files/Active Deck.xlsx')





def cardselection6():
    global wb2
    global ws2
    global rscreen
    global price
    global coinshopvar
    global purchacelimit
    if price > coinshopvar:
        print("Insufficient funds")
    else:
        purchacelimit=purchacelimit+1
        coinshopvar=coinshopvar-price
        coins.config(text=coinshopvar)
        if purchacelimit ==1:
            print("col 6 working")
            thercard = ws2['A12']
            thercard.value = "moreE"
            wb2.save('Excel files/Active Deck.xlsx')

        if purchacelimit ==2:
            print("col 6 working")
            thercard = ws2['A13']
            thercard.value = "moreE"
            wb2.save('Excel files/Active Deck.xlsx')

        if purchacelimit ==3:
            print("col 6 working")
            thercard = ws2['A14']
            thercard.value = "moreE"
            wb2.save('Excel files/Active Deck.xlsx')

        if purchacelimit ==4:
            print("col 6 working")
            thercard = ws2['A15']
            thercard.value = "moreE"
            wb2.save('Excel files/Active Deck.xlsx')



def cardselection4():
    global wb2
    global ws2
    global rscreen
    global price
    global coinshopvar
    global purchacelimit

    if price > coinshopvar:
        print("Insufficient funds")
    else:
        purchacelimit=purchacelimit+1
        coinshopvar=coinshopvar-price
        coins.config(text=coinshopvar)
    if purchacelimit ==1:
        print("col 4 working")
        thercard = ws2['A12']
        thercard.value = "double"
        wb2.save('Excel files/Active Deck.xlsx')
    if purchacelimit ==2:
        print("col 4 working")
        thercard = ws2['A13']
        thercard.value = "double"
        wb2.save('Excel files/Active Deck.xlsx')

    if purchacelimit ==3:
        print("col 4 working")
        thercard = ws2['A14']
        thercard.value = "double"
        wb2.save('Excel files/Active Deck.xlsx')

    if purchacelimit ==4:
        print("col 4 working")
        thercard = ws2['A15']
        thercard.value = "double"
        wb2.save('Excel files/Active Deck.xlsx')



def pricelisting7():
    global shopsn
    global price
    global botmidmultichoice
    global botmidrepairchoice
    global botmidbigblockchoice
    global botmidmoreEchoice
    global botmidheavyhitchoice
    global botmiddoublestrchoice
    global botmidcardleft
    global coinshopvar

    price=143

    if price > coinshopvar:
        print("still in stock")
    else:
        soldout = PhotoImage(file="images/soldout card.png")
        botmidcardleft.config(image=soldout)

    try:
        if botmidmultichoice ==True:
            cardselection1()
            botmidmultichoice = False
    except:
        print(".")
    try:
        if botmidrepairchoice ==True:
            cardselection2()
            botmidrepairchoice = False
    except:
        print(".")
    try:
        if botmidheavyhitchoice ==True:
            cardselection3()
            botmidheavyhitchoice = False
    except:
        print(".")
    try:
        if botmiddoublestrchoice ==True:
            cardselection4()
            botmiddoublestrchoice = False
    except:
        print(".")
    try:
        if botmidbigblockchoice ==True:
            cardselection5()
            botmidbigblockchoice = False
    except:
        print(".")
    try:
        if botmidmoreEchoice ==True:
            cardselection6()
            botmidmoreEchoice = False
    except:
        print(".")
    shopsn.mainloop()
def pricelisting6():
    global shopsn
    global price
    global botleftmultichoice
    global botleftrepairchoice
    global botleftbigblockchoice
    global botleftmoreEchoice
    global botleftheavyhitchoice
    global botleftdoublestrchoice
    global botleftcard
    global coinshopvar
    price=72

    if price > coinshopvar:
        print("still in stock")
    else:
        soldout=PhotoImage(file="images/soldout card.png")
        botleftcard.config(image=soldout)

    try:
        if botleftmultichoice ==True:
            cardselection1()
            botleftmultichoice = False
    except:
        print(".")
    try:
        if botleftrepairchoice ==True:
            cardselection2()
            botleftrepairchoice = False
    except:
        print(".")
    try:
        if botleftheavyhitchoice ==True:
            cardselection3()
            botleftheavyhitchoice = False
    except:
        print(".")
    try:
        if botleftdoublestrchoice ==True:
            cardselection4()
            botleftdoublestrchoice = False
    except:
        print(".")
    try:
        if botleftbigblockchoice ==True:
            cardselection5()
            botleftbigblockchoice = False
    except:
        print(".")
    try:
        if botleftmoreEchoice ==True:
            cardselection6()
            botleftmoreEchoice = False
    except:
        print(".")
    shopsn.mainloop()

def pricelisting5():
    global shopsn
    global price
    global rightmultichoice
    global rightrepairchoice
    global rightbigblockchoice
    global rightmoreEchoice
    global rightheavyhitchoice
    global rightdoublestrchoice
    global rightcard
    global coinshopvar


    price = 70
    if price > coinshopvar:
        print("still in stock")
    else:
        soldout = PhotoImage(file="images/soldout card.png")
        rightcard.config(image=soldout)
    try:
        if rightmultichoice == True:
            cardselection1()
            rightmultichoice = False
    except:
        print(".")
    try:
        if rightrepairchoice == True:
            cardselection2()
            rightrepairchoice = False
    except:
        print(".")
    try:
        if rightheavyhitchoice == True:
            cardselection3()
            rightheavyhitchoice = False
    except:
        print(".")
    try:
        if rightdoublestrchoice == True:
            cardselection4()
            rightdoublestrchoice = False
    except:
        print(".")
    try:
        if rightbigblockchoice == True:
            cardselection5()
            rightbigblockchoice = False
    except:
        print(".")
    try:
        if rightmoreEchoice == True:
            cardselection6()
            rightmoreEchoice = False
    except:
        print(".")
    shopsn.mainloop()

def pricelisting4():
    global shopsn
    global price
    global midrightmultichoice
    global midrightrepairchoice
    global midrightbigblockchoice
    global midrightmoreEchoice
    global midrightheavyhitchoice
    global midrightdoublestrchoice
    global midcardright
    global coinshopvar



    price = 69
    if price > coinshopvar:
        print("still in stock")
    else:
        soldout=PhotoImage(file="images/soldout card.png")
        midcardright.config(image=soldout)

    try:
        if midrightmultichoice == True:
            cardselection1()
            midrightmultichoice = False
    except:
        print(".")
    try:
        if midrightrepairchoice == True:
            cardselection2()
            midrightrepairchoice = False
    except:
        print(".")
    try:
        if midrightheavyhitchoice == True:
            cardselection3()
            midrightheavyhitchoice = False
    except:
        print(".")
    try:
        if midrightdoublestrchoice == True:
            cardselection4()
            midrightdoublestrchoice = False
    except:
        print(".")
    try:
        if midrightbigblockchoice == True:
            cardselection5()
            midrightbigblockchoice = False
    except:
        print(".")
    try:
        if midrightmoreEchoice == True:
            cardselection6()
            midrightmoreEchoice = False
    except:
        print(".")
    shopsn.mainloop()

def pricelisting3():
    global shopsn
    global price
    global midmultichoice
    global midrepairchoice
    global midbigblockchoice
    global midmoreEchoice
    global midheavyhitchoice
    global middoublestrchoice
    global centralcard
    global coinshopvar


    price = 46
    if price > coinshopvar:
        print("still in stock")
    else:
        soldout=PhotoImage(file="images/soldout card.png")
        centralcard.config(image=soldout)
    try:
        if midmultichoice == True:
            cardselection1()
            midmultichoice = False
    except:
        print(".")
    try:
        if midrepairchoice == True:
            cardselection2()
            midrepairchoice = False
    except:
        print(".")
    try:
        if midheavyhitchoice == True:
            cardselection3()
            midheavyhitchoice = False
    except:
        print(".")
    try:
        if middoublestrchoice == True:
            cardselection4()
            middoublestrchoice = False
    except:
        print(".")
    try:
        if midbigblockchoice == True:
            cardselection5()
            midbigblockchoice = False
    except:
        print(".")
    try:
        if midmoreEchoice == True:
            cardselection6()
            midmoreEchoice = False
    except:
        print(".")
    shopsn.mainloop()
def pricelisting2():
    global shopsn
    global price
    global midleftmultichoice
    global midleftrepairchoice
    global midleftbigblockchoice
    global midleftmoreEchoice
    global midleftheavyhitchoice
    global midleftdoublestrchoice
    global midcardleft
    global coinshopvar


    price = 46
    if price > coinshopvar:
        print("still in stock")
    else:
        soldout=PhotoImage(file="images/soldout card.png")
        midcardleft.config(image=soldout)
    try:
        if midleftmultichoice == True:
            cardselection1()
            midleftmultichoice = False
    except:
        print(".")
    try:
        if midleftrepairchoice == True:
            cardselection2()
            midleftrepairchoice = False
    except:
        print(".")
    try:
        if midleftheavyhitchoice == True:
            cardselection3()
            midleftheavyhitchoice = False
    except:
        print(".")

    try:
        if midleftdoublestrchoice == True:
            cardselection4()
            midleftdoublestrchoice = False
    except:
        print(".")
    try:
        if midleftbigblockchoice == True:
            cardselection5()
            midleftbigblockchoice = False
    except:
        print(".")
    try:
        if midleftmoreEchoice == True:
            cardselection6()
            midleftmoreEchoice = False
    except:
        print(".")
    shopsn.mainloop()
def pricelisting1():
    global shopsn
    global price
    global leftmultichoice
    global leftrepairchoice
    global leftbigblockchoice
    global leftmoreEchoice
    global leftheavyhitchoice
    global leftdoublestrchoice
    global leftcard
    global coinshopvar

    price=68
    if price > coinshopvar:
        print("still in stock")
    else:
        soldout=PhotoImage(file="images/soldout card.png")
        leftcard.config(image=soldout)
    try:
        if leftmultichoice ==True:
            cardselection1()
            leftmultichoice = False
    except:
        print(".")

    try:
        if leftrepairchoice ==True:
            cardselection2()
            leftrepairchoice = False
    except:
        print(".")

    try:
        if leftheavyhitchoice ==True:
            cardselection3()
            leftheavyhitchoice = False
    except:
        print(".")

    try:
        if leftdoublestrchoice ==True:
            cardselection4()
            leftdoublestrchoice = False
    except:
        print(".")
    try:
        if leftbigblockchoice ==True:
            cardselection5()
            leftbigblockchoice = False
    except:
        print(".")
    try:
        if leftmoreEchoice ==True:
            cardselection6()
            leftmoreEchoice = False
    except:
        print(".")
    shopsn.mainloop()


def cardoptions():
    global shopsn
    global price
    global leftcard
    global rightcard
    global midcard
    global midcardleft
    global midcardright
    global centralcard
    global botleftcard
    global botmidcardleft

    global leftmultichoice
    global leftrepairchoice
    global leftbigblockchoice
    global leftmoreEchoice
    global leftheavyhitchoice
    global leftdoublestrchoice

    global midleftmultichoice
    global midleftrepairchoice
    global midleftbigblockchoice
    global midleftmoreEchoice
    global midleftheavyhitchoice
    global midleftdoublestrchoice

    global midmultichoice
    global midrepairchoice
    global midbigblockchoice
    global midmoreEchoice
    global midheavyhitchoice
    global middoublestrchoice

    global midrightmultichoice
    global midrightrepairchoice
    global midrightbigblockchoice
    global midrightmoreEchoice
    global midrightheavyhitchoice
    global midrightdoublestrchoice

    global rightmultichoice
    global rightrepairchoice
    global rightbigblockchoice
    global rightmoreEchoice
    global rightheavyhitchoice
    global rightdoublestrchoice

    global botleftmultichoice
    global botleftrepairchoice
    global botleftbigblockchoice
    global botleftmoreEchoice
    global botleftheavyhitchoice
    global botleftdoublestrchoice

    global botmidmultichoice
    global botmidrepairchoice
    global botmidbigblockchoice
    global botmidmoreEchoice
    global botmidheavyhitchoice
    global botmiddoublestrchoice


    leftcardchoice=random.randint(1,6)
    botleftcardchoice = random.randint(1, 6)
    midcardleftchoice=random.randint(1,6)
    botmidcardleftchoice=random.randint(1,6)
    centralcardchoice=random.randint(1,6)
    midcardrightchoice=random.randint(1,6)
    rightcardchoice=random.randint(1,6)

    bigblock=PhotoImage(file="images/30 block card.png")
    repair=PhotoImage(file="images/Self repair card.png")
    heavyhit=PhotoImage(file="images/Heavy hit card reskin.png")
    multihit=PhotoImage(file="images/multi slash card.png")
    doublestr=PhotoImage(file="images/Double Strength card.png")
    more_energy=PhotoImage(file="images/Double Energy.png")

    if leftcardchoice==1:
        leftcard=Button(shopsn,image=multihit,command=pricelisting1)
        leftcard.place(relx=0.175,y=120)
        leftmultichoice=True

    if leftcardchoice==2:
        leftcard=Button(shopsn,image=repair,command=pricelisting1)
        leftcard.place(relx=0.175,y=120)
        leftrepairchoice=True

    if leftcardchoice == 3:
        leftcard = Button(shopsn, image=heavyhit,command=pricelisting1)
        leftcard.place(relx=0.175, y=120)
        leftheavyhitchoice=True


    if leftcardchoice == 4:
        leftcard = Button(shopsn, image=doublestr,command=pricelisting1)
        leftcard.place(relx=0.175, y=120)
        leftdoublestrchoice=True

    if leftcardchoice == 5:
        leftcard = Button(shopsn, image=bigblock,command=pricelisting1)
        leftcard.place(relx=0.175, y=120)
        leftbigblockchoice=True


    if leftcardchoice == 6:
        leftcard = Button(shopsn, image=more_energy,command=pricelisting1)
        leftcard.place(relx=0.175, y=120)
        leftmoreEchoice=True

    if midcardleftchoice == 1:
        midcardleft = Button(shopsn, image=multihit,command=pricelisting2)
        midcardleft.place(relx=0.32, y=120)
        midleftmultichoice = True

    if midcardleftchoice == 2:
        midcardleft = Button(shopsn, image=repair,command=pricelisting2)
        midcardleft.place(relx=0.32, y=120)
        midleftrepairchoice = True


    if midcardleftchoice == 3:
        midcardleft = Button(shopsn, image=heavyhit,command=pricelisting2)
        midcardleft.place(relx=0.32, y=120)
        midleftheavyhitchoice = True

    if midcardleftchoice == 4:
        midcardleft = Button(shopsn, image=doublestr,command=pricelisting2)
        midcardleft.place(relx=0.32, y=120)
        midleftdoublestrchoice = True

    if midcardleftchoice == 5:
        midcardleft = Button(shopsn, image=bigblock,command=pricelisting2)
        midcardleft.place(relx=0.32, y=120)
        midleftbigblockchoice = True

    if midcardleftchoice == 6:
        midcardleft = Button(shopsn, image=more_energy,command=pricelisting2)
        midcardleft.place(relx=0.32, y=120)
        midleftmoreEchoice = True

    if centralcardchoice == 1:
        centralcard = Button(shopsn, image=multihit,command=pricelisting3)
        centralcard.place(relx=0.465, y=120)
        midmultichoice = True

    if centralcardchoice == 2:
        centralcard = Button(shopsn, image=repair,command=pricelisting3)
        centralcard.place(relx=0.465, y=120)
        midrepairchoice = True

    if centralcardchoice == 3:
        centralcard = Button(shopsn, image=heavyhit,command=pricelisting3)
        centralcard.place(relx=0.465, y=120)
        midheavyhitchoice = True

    if centralcardchoice == 4:
        centralcard = Button(shopsn, image=doublestr,command=pricelisting3)
        centralcard.place(relx=0.465, y=120)
        middoublestrchoice = True

    if centralcardchoice == 5:
        centralcard = Button(shopsn, image=bigblock,command=pricelisting3)
        centralcard.place(relx=0.465, y=120)
        midbigblockchoice = True

    if centralcardchoice == 6:
        centralcard = Button(shopsn, image=more_energy,command=pricelisting3)
        centralcard.place(relx=0.465, y=120)
        midmoreEchoice = True

    if midcardrightchoice == 1:
        midcardright = Button(shopsn, image=multihit,command=pricelisting4)
        midcardright.place(relx=0.615, y=120)
        midrightmultichoice = True

    if midcardrightchoice == 2:
        midcardright = Button(shopsn, image=repair,command=pricelisting4)
        midcardright.place(relx=0.615, y=120)
        midrightrepairchoice = True

    if midcardrightchoice == 3:
        midcardright = Button(shopsn, image=heavyhit,command=pricelisting4)
        midcardright.place(relx=0.615, y=120)
        midrightheavyhitchoice = True

    if midcardrightchoice == 4:
        midcardright = Button(shopsn, image=doublestr,command=pricelisting4)
        midcardright.place(relx=0.615, y=120)
        midrightdoublestrchoice = True

    if midcardrightchoice == 5:
        midcardright = Button(shopsn, image=bigblock,command=pricelisting4)
        midcardright.place(relx=0.615, y=120)
        midrightbigblockchoice = True

    if midcardrightchoice == 6:
        midcardright = Button(shopsn, image=more_energy,command=pricelisting4)
        midcardright.place(relx=0.615, y=120)
        midrightmoreEchoice = True

    if rightcardchoice == 1:
        rightcard = Button(shopsn, image=multihit, command=pricelisting5)
        rightcard.place(relx=0.755, y=120)
        rightmultichoice = True

    if rightcardchoice == 2:
        rightcard = Button(shopsn, image=repair, command=pricelisting5)
        rightcard.place(relx=0.755, y=120)
        rightrepairchoice = True

    if rightcardchoice == 3:
        rightcard = Button(shopsn, image=heavyhit, command=pricelisting5)
        rightcard.place(relx=0.755, y=120)
        rightheavyhitchoice = True

    if rightcardchoice == 4:
        rightcard = Button(shopsn, image=doublestr, command=pricelisting5)
        rightcard.place(relx=0.755, y=120)
        rightdoublestrchoice = True

    if rightcardchoice == 5:
        rightcard = Button(shopsn, image=bigblock, command=pricelisting5)
        rightcard.place(relx=0.755, y=120)
        rightbigblockchoice = True

    if rightcardchoice == 6:
        rightcard = Button(shopsn, image=more_energy, command=pricelisting5)
        rightcard.place(relx=0.755, y=120)
        rightmoreEchoice = True

    if botleftcardchoice==1:
        botleftcard=Button(shopsn,image=multihit,command=pricelisting6)
        botleftcard.place(relx=0.175,y=450)
        botleftmultichoice = True

    if botleftcardchoice==2:
        botleftcard=Button(shopsn,image=repair,command=pricelisting6)
        botleftcard.place(relx=0.175,y=450)
        botleftrepairchoice = True

    if botleftcardchoice == 3:
        botleftcard = Button(shopsn, image=heavyhit,command=pricelisting6)
        botleftcard.place(relx=0.175, y=450)
        botleftheavyhitchoice = True

    if botleftcardchoice == 4:
        botleftcard = Button(shopsn, image=doublestr,command=pricelisting6)
        botleftcard.place(relx=0.175, y=450)
        botleftdoublestrchoice = True

    if botleftcardchoice == 5:
        botleftcard = Button(shopsn, image=bigblock,command=pricelisting6)
        botleftcard.place(relx=0.175, y=450)
        botleftbigblockchoice = True

    if botleftcardchoice == 6:
        botleftcard = Button(shopsn, image=more_energy,command=pricelisting6)
        botleftcard.place(relx=0.175, y=450)
        botleftmoreEchoice = True

    if botmidcardleftchoice == 1:
        botmidcardleft = Button(shopsn, image=multihit,command=pricelisting7)
        botmidcardleft.place(relx=0.32, y=450)
        botmidmultichoice = True

    if botmidcardleftchoice == 2:
        botmidcardleft = Button(shopsn, image=repair,command=pricelisting7)
        botmidcardleft.place(relx=0.32, y=450)
        botmidrepairchoice = True

    if botmidcardleftchoice == 3:
        botmidcardleft = Button(shopsn, image=heavyhit,command=pricelisting7)
        botmidcardleft.place(relx=0.32, y=450)
        botmidheavyhitchoice = True

    if botmidcardleftchoice == 4:
        botmidcardleft = Button(shopsn, image=doublestr,command=pricelisting7)
        botmidcardleft.place(relx=0.32, y=450)
        botmiddoublestrchoice = True

    if botmidcardleftchoice == 5:
        botmidcardleft = Button(shopsn, image=bigblock,command=pricelisting7)
        botmidcardleft.place(relx=0.32, y=450)
        botmidbigblockchoice = True

    if botmidcardleftchoice == 6:
        botmidcardleft = Button(shopsn, image=more_energy,command=pricelisting7)
        botmidcardleft.place(relx=0.32, y=450)
        botmidmoreEchoice = True

    shopsn.mainloop()

def Displayoverlay():
    global mainguy
    global shopkeeper
    global entershopbutton
    global shopsn
    global coinbaglabel
    global coins
    global coinshopvar
    coincolourbg = '#%02x%02x%02x' % (61,73,79)
    coinbaglabel.place_forget()
    coins.place_forget()
    mainguy.place_forget()
    shopguy.place_forget()
    entershopbutton.place_forget()

    shopoverlay=PhotoImage(file="images/shop_overlay.png")
    shopoverlayLabel = Label(shopsn,image=shopoverlay,anchor="center")
    shopoverlayLabel.place(relx=0,y=0)

    cardremovalimage=PhotoImage(file="images/card_removal.png")
    cardremoval=Button(shopsn,image=cardremovalimage,command=cardRemover)
    cardremoval.place(relx=0.75,y=448)

    coinbag = PhotoImage(file="images/coinbag.png")
    coinbaglabel = Label(shopsn, image=coinbag)
    coinbaglabel.place(relx=0, y=0)
    coins = Label(shopsn, text="195", font=("Times New Roman", 30,), bg=coincolourbg, fg="yellow")
    coins.place(relx=0.05, y=2)
    coinshopvar=195
    leavephoto=PhotoImage(file="images/shop_leavebutton.png")
    leavebutton=Button(shopsn,image=leavephoto,command=exit)
    leavebutton.place(relx=0,y=695)
    cardoptions()
    shopsn.mainloop()

class ImageLabel(Label):
    """a label that displays images, and plays them if they are gifs"""
    def load(self, im):
        if isinstance(im, str):
            im = Image.open(im)
        self.loc = 0
        self.frames = []

        try:
            for i in count(1):
                self.frames.append(ImageTk.PhotoImage(im.copy()))
                im.seek(i)
        except EOFError:
            pass

        try:
            self.delay = im.info['duration']
        except:
            self.delay = 100

        if len(self.frames) == 1:
            self.config(image=self.frames[0])
        else:
            self.next_frame()

    def unload(self):
        self.config(image="")
        self.frames = None

    def next_frame(self):
        if self.frames:
            self.loc += 1
            self.loc %= len(self.frames)
            self.config(image=self.frames[self.loc])
            self.after(self.delay, self.next_frame)



def entershop():
    global shopsn
    global mainguy
    global shopguy
    global entershopbutton
    global coinbaglabel
    global coins
    global ws2
    global wb2
    global price
    global purchacelimit
    global psL
    global pbL
    shopsn=Toplevel()
    shopsn.geometry("1500x805")
    bg = PhotoImage(file="images/shop.png")
    bglabel=Label(shopsn,image=bg)
    bglabel.place(relx=0,y=0)
    wb2 = load_workbook("Excel files/Active Deck.xlsx")
    ws2=wb2.active
    wcL = load_workbook("Excel files/Character Selection.xlsx")
    wcS = wcL.active
    pbL = load_workbook("Excel files/Player Data.xlsx")
    psL=pbL.active
    coincolourbg = '#%02x%02x%02x' % (61,73,79)
    price=0
    purchacelimit=0
    coinbag = PhotoImage(file="images/coinbag.png")
    coinbaglabel = Label(shopsn, image=coinbag)
    coinbaglabel.place(relx=0, y=0)
    coins = Label(shopsn, text="195", font=("Times New Roman", 30,), bg=coincolourbg, fg="yellow")
    coins.place(relx=0.05, y=2)

    mainguy=ImageLabel(shopsn)
    mainguy.place(relx=0.2, y=500, anchor="center")
    if wcS['A1'].value == "Defect":
        mainguy.load('images/The Defect.gif')
    if wcS['A1'].value == "Silent":
        mainguy.load("images/thief.gif")
    if wcS['A1'].value == "Ironclad":
        mainguy.load("images/Ironclad.gif")

    shopguy=ImageLabel(shopsn)
    shopguy.place(relx=0.67, y=390, anchor="center")
    shopguy.load("images/shopkeeper.gif")

    esb=PhotoImage(file="images/entershop button.png")
    entershopbutton=Button(shopsn,image=esb,anchor="center",command=Displayoverlay)
    entershopbutton.place(relx=0.5,y=500)



    shopsn.mainloop()


