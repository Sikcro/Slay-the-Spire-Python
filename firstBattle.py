from tkinter import *
from openpyxl import workbook, load_workbook
import random
import pygame
from PIL import Image, ImageTk
from itertools import count
from RewardScreen import Rewardstart
from Death import PLayerDied

def gameover():
    global firstbscreen
    firstbscreen.destroy()

    PLayerDied()



def RewardScreen():
    global firstbscreen
    global pbL
    global psL
    global healthvar
    psL['B2'].value = healthvar
    pbL.save('Excel files/Player Data.xlsx')
    firstbscreen.destroy()

    Rewardstart()



def enemydefend():
    global health
    global healthlabel
    global healthvar
    global enemyhasdefence
    global enemyshieldbar
    global enemyshieldvar

    enemyhasdefence=True
    shieldbarpng=PhotoImage(file="images/shield.png")
    enemyshieldbar=Label(firstbscreen,image=shieldbarpng)
    enemyshieldbar.place(relx=0.73,y=530)
    Eshield=6
    enemyshieldvar=Label(firstbscreen,text=Eshield)
    enemyshieldvar.place(relx=0.72,y=530)
    firstbscreen.mainloop()




def enemyatk():
    global health
    global healthlabel
    global healthvar
    global Efirstatkcomplete
    global Esecondatkcomplete
    global Ethirdatkcomplete
    global Efourthatkcomplete
    global Efithatkcomplete
    global Esixthatkcomplete
    global Eeighthatkcomplete
    global Eseventhatkcomplete
    global enemyatksetup
    global shieldactivated
    global Shieldvar
    global shieldlabel
    global shieldnum
    seventypercent=PhotoImage(file="images/healthbar 70 percent dmg.png")
    ninetypercent=PhotoImage(file="images/healthbar 90 percent dmg.png")
    fiftypercent=PhotoImage(file="images/healthbar half dmg.png")
    twentypercent=PhotoImage(file="images/healthbar small dmg.png")
    onehundredpercent=PhotoImage(file="images/healthbar 100 percent dmg.png")
    enemyatkpower=10

    try:
        if shieldactivated == True:
            enemyatkpower=enemyatkpower-Shieldvar
            Shieldvar=0
            shieldnum.destroy()
            shieldlabel.destroy()

        if enemyatkpower<0:
            enemyatkpower=0
            print("working")
        else:
            print("enemy pwr not less than 0")
        healthvar=healthvar-enemyatkpower

    except:
        print("no shield played this turn")




    if healthvar <=0:
        gameover()

    else:

        try:
            if Eseventhatkcomplete == True:

                if shieldactivated == True:
                    healthlabel.configure(text=(healthvar, "/75"))
                    shieldactivated=False
                else:
                    health.configure(image=onehundredpercent)
                    healthlabel.configure(text=(healthvar,"/75"))
                    Eseventhatkcomplete = False
                    print("YOU lost")
                    gameover()

        except:
            print(".")




        try:
            if Esixthatkcomplete == True:
                if shieldactivated == True:
                    healthlabel.configure(text=(healthvar, "/75"))
                    shieldactivated=False

                else:
                    health.configure(image=onehundredpercent)
                    healthlabel.configure(text=(healthvar,"/75"))
                    Esixthatkcomplete = False
                    Eseventhatkcomplete = True
        except:
            print(".")



        try:
            if Efithatkcomplete == True:
                if shieldactivated == True:
                    healthlabel.configure(text=(healthvar, "/75"))
                    shieldactivated=False

                else:
                    health.configure(image=ninetypercent)
                    healthlabel.configure(text=(healthvar,"/75"))
                    Efithatkcomplete = False
                    Esixthatkcomplete= True

        except:
            print(".")


        try:
            if Efourthatkcomplete==True:

                if shieldactivated == True:
                    healthlabel.configure(text=(healthvar, "/75"))
                    shieldactivated=False
                else:
                    health.configure(image=ninetypercent)
                    healthlabel.configure(text=(healthvar,"/75"))
                    Efourthatkcomplete=False
                    Efithatkcomplete=True

        except:
            print(".")



        try:
            if Ethirdatkcomplete==True:

                if shieldactivated == True:
                    healthlabel.configure(text=(healthvar, "/75"))
                    shieldactivated=False
                else:
                    health.configure(image=ninetypercent)
                    healthlabel.configure(text=(healthvar,"/75"))
                    Ethirdatkcomplete=False
                    Efourthatkcomplete=True
        except:
            print(".")


        try:
            if Esecondatkcomplete==True:
                if shieldactivated == True:
                    healthlabel.configure(text=(healthvar, "/75"))
                    shieldactivated=False
                else:
                    health.configure(image=seventypercent)
                    healthlabel.configure(text=(healthvar,"/75"))
                    Ethirdatkcomplete=True
        except:
            print(".")

        try:
            if Efirstatkcomplete==True:

                if shieldactivated == True:
                    healthlabel.configure(text=(healthvar, "/75"))
                    shieldactivated=False
                else:
                    health.configure(image=fiftypercent)
                    healthlabel.configure(text=(healthvar,"/75"))
                    Efirstatkcomplete=False
                    Esecondatkcomplete=True
        except:
            print(".")

        if  enemyatksetup==True:

            if shieldactivated==True:
                healthlabel.configure(text=(healthvar, "/75"))
                shieldactivated=False

            else:
                health.configure(image=twentypercent)
                healthlabel.configure(text=(healthvar,"/75"))
                Efirstatkcomplete=True
                enemyatksetup=False

        else:
            print(".")

    firstbscreen.mainloop()



def enemydecision():
    global enemyintentlabel
    enemydice=random.randint(1,1)
    atkintent = PhotoImage(file="images/atk intent.png")
    defintent = PhotoImage(file="images/shield intent.PNG")
    if enemydice ==1:
        enemyintentlabel.config(image=atkintent)
        enemyatk()

    if enemydice ==2:
        enemyintentlabel.config(image=defintent)
        enemydefend()

    firstbscreen.mainloop()

def defencecard():
    global firstbscreen
    global Shieldvar
    global shieldactivated
    global shieldlabel
    global shieldnum
    global cardclicked
    shieldactivated=True
    shieldimg=PhotoImage(file="images/shield.png")
    cardclicked=cardclicked+1

    if energynumber <0:
        print("cannot apply Block")

    else:
        pygame.mixer.init()
        pygame.mixer.music.load("Audio/r_laser_p-95552.mp3")
        pygame.mixer.music.play(loops=1)

        if cardclicked==1:
            try:
                shieldlabel.place_forget()
                shieldnum.place_forget()
            except:
                print("-")
            shieldlabel=Label(firstbscreen,image=shieldimg)
            shieldlabel.place(relx=0.135, y=630)
            Shieldvar=Shieldvar+5
            shieldnum=Label(firstbscreen,text=Shieldvar)
            shieldnum.place(relx=0.135, y=630)
        if cardclicked==2:
            try:
                shieldlabel.place_forget()
                shieldnum.place_forget()
            except:
                print("-")
            shieldlabel=Label(firstbscreen,image=shieldimg)
            shieldlabel.place(relx=0.135, y=630)
            Shieldvar=Shieldvar+5
            shieldnum=Label(firstbscreen,text=Shieldvar)
            shieldnum.place(relx=0.135, y=630)

        if cardclicked==3:
            try:
                shieldlabel.place_forget()
                shieldnum.place_forget()
            except:
                print("-")

            shieldlabel=Label(firstbscreen,image=shieldimg)
            shieldlabel.place(relx=0.135, y=630)
            Shieldvar=Shieldvar+5
            shieldnum=Label(firstbscreen,text=Shieldvar)
            shieldnum.place(relx=0.135, y=630)
            cardclicked=0

    firstbscreen.mainloop()


def strikecard():
    global enemyhealthlabel
    global enemyhealth
    global enemyhealthvar
    global firstbscreen
    global atksetup
    global firstatkcomplete
    global secondatkcomplete
    global thirdatkcomplete
    global fourthatkcomplete
    global bg
    global enemyhasdefence
    global enemyshieldvar
    global enemyshieldbar
    global energynumber
    global pbL
    global psL


    seventypercent=PhotoImage(file="images/healthbar 70 percent dmg.png")
    ninetypercent=PhotoImage(file="images/healthbar 90 percent dmg.png")
    fiftypercent=PhotoImage(file="images/healthbar half dmg.png")
    twentypercent=PhotoImage(file="images/healthbar small dmg.png")
    onehundredpercent=PhotoImage(file="images/healthbar 100 percent dmg.png")
    reward=PhotoImage(file="images/win screen.png")
    attackpower=6

    if energynumber <0:
        print("cannot attack")

    else:
        pygame.mixer.init()
        pygame.mixer.music.load("Audio/cartoon punch 1.mp3")
        pygame.mixer.music.play(loops=1)
        enemyhealthvar=enemyhealthvar-attackpower

        try:
            if enemyhasdefence==True:
                enemyshieldvar.place_forget()
                enemyshieldbar.place_forget()



        except:
            print("enemy has no shield")



        try:
            if fourthatkcomplete==True:

                if enemyhasdefence==True:
                    enemyhealthvar = enemyhealthvar + attackpower
                    enemyhealth.configure(image=ninetypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    fourthatkcomplete = True
                    enemyhasdefence=False

                else:
                    enemyhealth.configure(image=onehundredpercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)

                    fourthatkcomplete=False
                    print("YOU won!!!")
                    coinvalue=psL['B1']
                    coinvalue.value=75
                    pbL.save('Excel files/Player Data.xlsx')

                    RewardScreen()




        except:
            print(".")



        try:
            if thirdatkcomplete==True:

                if enemyhasdefence==True:
                    enemyhealthvar = enemyhealthvar + attackpower
                    enemyhealth.configure(image=seventypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    thirdatkcomplete=True
                    fourthatkcomplete=False
                    enemyhasdefence=False
                else:
                    from TheWorld import coinslvl1clear
                    enemyhealth.configure(image=ninetypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    thirdatkcomplete=False
                    fourthatkcomplete=True
                    firstbscreen.after(100,coinslvl1clear)


        except:
            print(".")


        try:
            if secondatkcomplete==True:

                if enemyhasdefence==True:
                    enemyhealthvar = enemyhealthvar + attackpower
                    enemyhealth.configure(image=fiftypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    secondatkcomplete=True
                    thirdatkcomplete=False
                    enemyhasdefence=False
                else:
                    enemyhealth.configure(image=seventypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    secondatkcomplete=False
                    thirdatkcomplete=True

        except:
            print(".")

        try:
            if firstatkcomplete==True:

                if enemyhasdefence==True:
                    enemyhealthvar=enemyhealthvar+attackpower
                    enemyhealth.configure(image=twentypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    secondatkcomplete=False
                    firstatkcomplete=True
                    enemyhasdefence=False
                else:
                    enemyhealth.configure(image=fiftypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    firstatkcomplete=False
                    secondatkcomplete=True

        except:
            print(".")

        if  atksetup==True:
            enemyhealth.configure(image=twentypercent)
            enemyhealthlabel.configure(text=enemyhealthvar)
            firstatkcomplete=True
            atksetup=False


        else:
            print(".")


    firstbscreen.mainloop()







def deckcreation():
    global atkcardcounter
    global blockcardcounter
    global Drawlabel
    decktotal=atkcardcounter+blockcardcounter

    if decktotal <0:
        decktotal=3

    Drawlabel.configure(text=decktotal)




def discardpile5():
    global slot5
    global Energylabel
    global discardlabel
    global energynumber
    global discardnumber
    global slot5cardplayed

    slot5cardplayed=True
    if energynumber > 0:
        slot5.place_forget()
        energynumber=energynumber-1
        print("your energy number is", energynumber)
        discardnumber=discardnumber+1
        Energylabel.configure(text=(energynumber,"/3"))
        discardlabel.configure(text=discardnumber)
    else:
        energynumber=-5
        print("you got no energy!!!")



def discardpile4():
    global slot4
    global Energylabel
    global discardlabel
    global energynumber
    global discardnumber
    global slot4cardplayed

    slot4cardplayed=True
    if energynumber > 0:
        slot4.place_forget()
        energynumber=energynumber-1
        print("your energy number is",energynumber)
        discardnumber=discardnumber+1
        Energylabel.configure(text=(energynumber,"/3"))
        discardlabel.configure(text=discardnumber)
    else:
        energynumber = -5
        print("you got no energy !!!")



def discardpile3():
    global slot3
    global Energylabel
    global discardlabel
    global energynumber
    global discardnumber
    global slot3cardplayed

    slot3cardplayed=True
    if energynumber > 0:
        slot3.place_forget()
        energynumber=energynumber-1
        print("your energy number is", energynumber)
        discardnumber=discardnumber+1
        Energylabel.configure(text=(energynumber,"/3"))
        discardlabel.configure(text=discardnumber)
    else:
        energynumber = -5
        print("you got no energy!!!")


def discardpile2():
    global slot2
    global Energylabel
    global discardlabel
    global energynumber
    global discardnumber
    global slot2cardplayed


    slot2cardplayed=True
    if energynumber > 0:
        slot2.place_forget()
        energynumber=energynumber-1
        print("your energy number is", energynumber)
        discardnumber=discardnumber+1
        Energylabel.configure(text=(energynumber,"/3"))
        discardlabel.configure(text=discardnumber)
    else:
        energynumber = -5
        print("you got not energy!!!")



def discardpile():
    global slot1
    global Energylabel
    global discardlabel
    global energynumber
    global discardnumber
    global slot1cardplayed

    slot1cardplayed=True
    if energynumber >0:
        slot1.place_forget() # first card wipe process ------
        energynumber=energynumber-1

        discardnumber=discardnumber+1


        Energylabel.configure(text=(energynumber,"/3"))
        discardlabel.configure(text=discardnumber)
    else:
        energynumber = -5
        print("you got no Energy!!!")






def auto_drawhand_animslot5():
    global str
    global firstbscreen
    global atkcardcounter
    global blockcardcounter
    global blo
    global slot5

    diceslot5 = random.randint(1, 2)

    if diceslot5 == 1:


        if wsL['A2'].value =="Reskin":
            slot5 = Button(firstbscreen, image=strEX, command=lambda: [discardpile5(), strikecard()])
            slot5.place(relx=0.6, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

        else:
            slot5 = Button(firstbscreen, image=str,command=lambda:[discardpile5(),strikecard()])
            slot5.place(relx=0.6, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

    if diceslot5 == 2:
        if wsL['A2'].value =="Reskin":
            slot5 = Button(firstbscreen, image=bloEX, command=lambda: [discardpile5(), defencecard()])
            slot5.place(relx=0.6, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

        else:
            slot5 = Button(firstbscreen, image=blo,command=lambda:[discardpile5(),defencecard()])
            slot5.place(relx=0.6, y=800, anchor="center")
            blockcardcounter = blockcardcounter - 1

    print("atk card in deck now are",atkcardcounter)
    print("block cards in deck now are",blockcardcounter)
    deckcreation()
    firstbscreen.mainloop()


def auto_drawhand_animslot4():
    global str
    global blo
    global firstbscreen
    global atkcardcounter
    global blockcardcounter
    global slot4
    diceslot4 = random.randint(1, 2)

    if diceslot4 == 1:

        if wsL['A2'].value =="Reskin":
            slot4 = Button(firstbscreen, image=strEX, command=lambda: [discardpile4(), strikecard()])
            slot4.place(relx=0.5, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

        else:
            slot4 = Button(firstbscreen, image=str,command=lambda:[discardpile4(),strikecard()])
            slot4.place(relx=0.5, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

    if diceslot4 == 2:

        if wsL['A2'].value =="Reskin":
            slot4 = Button(firstbscreen, image=bloEX, command=lambda: [discardpile4(), defencecard()])
            slot4.place(relx=0.5, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

        else:
            slot4 = Button(firstbscreen, image=blo,command=lambda:[discardpile4(),defencecard()])
            slot4.place(relx=0.5, y=800, anchor="center")
            blockcardcounter = blockcardcounter - 1
    firstbscreen.mainloop()


def auto_drawhand_animslot3():
    global str
    global blo
    global firstbscreen
    global atkcardcounter
    global blockcardcounter
    global slot3
    diceslot3 = random.randint(1,2)

    if diceslot3 == 1:

        if wsL['A2'].value =="Reskin":
            slot3 = Button(firstbscreen, image=strEX, command=lambda: [discardpile3(), strikecard()])
            slot3.place(relx=0.4, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

        else:
            slot3 = Button(firstbscreen, image=str,command=lambda:[discardpile3(),strikecard()])
            slot3.place(relx=0.4, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

    if diceslot3 == 2:
        if wsL['A2'].value =="Reskin":
            slot3 = Button(firstbscreen, image=bloEX, command=lambda: [discardpile3(), defencecard()])
            slot3.place(relx=0.4, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1
        else:
            slot3 = Button(firstbscreen, image=blo,command=lambda:[discardpile3(),defencecard()])
            slot3.place(relx=0.4, y=800, anchor="center")
            blockcardcounter = blockcardcounter - 1

    firstbscreen.mainloop()

def auto_drawhand_animslot2():
    global str
    global firstbscreen
    global blo
    global atkcardcounter
    global blockcardcounter
    global slot2
    diceslot2 = random.randint(1, 2)

    if diceslot2 == 1:
        if wsL['A2'].value =="Reskin":
            slot2 = Button(firstbscreen, image=strEX, command=lambda: [discardpile2(), strikecard()])
            slot2.place(relx=0.3, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1
        else:
            slot2 = Button(firstbscreen, image=str,command=lambda:[discardpile2(),strikecard()])
            slot2.place(relx=0.3, y=800, anchor="center")
            atkcardcounter = atkcardcounter-1


    if diceslot2 == 2:

        if wsL['A2'].value =="Reskin":
            slot2 = Button(firstbscreen, image=bloEX, command=lambda: [discardpile2(), defencecard()])
            slot2.place(relx=0.3, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

        else:
            slot2 = Button(firstbscreen, image=blo,command=lambda:[discardpile2(),defencecard()])
            slot2.place(relx=0.3, y=800, anchor="center")
            blockcardcounter = blockcardcounter-1


    firstbscreen.mainloop()

def auto_drawhand_animslot1():
    global str
    global firstbscreen
    global blo
    global atkcardcounter
    global blockcardcounter
    global slot1
    global wsL
    global strEX
    diceslot1=random.randint(1,2)


    if diceslot1 ==1:

        if wsL['A2'].value =="Reskin":
            slot1 = Button(firstbscreen, image=strEX, command=lambda: [discardpile(), strikecard()])
            slot1.place(relx=0.2, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

        else:
            slot1 = Button(firstbscreen, image=str,command=lambda:[discardpile(),strikecard()])
            slot1.place(relx=0.2, y=800, anchor="center")
            atkcardcounter=atkcardcounter-1




    if diceslot1 ==2:

        if wsL['A2'].value =="Reskin":
            slot1 = Button(firstbscreen, image=bloEX, command=lambda: [discardpile(), defencecard()])
            slot1.place(relx=0.2, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

        else:
            slot1 = Button(firstbscreen, image=blo,command=lambda:[discardpile(),defencecard()])
            slot1.place(relx=0.2, y=800, anchor="center")
            blockcardcounter=blockcardcounter-1




    firstbscreen.mainloop()

def autodrawhand():
    global ws2
    global firstbscreen
    global anim2
    global str
    global blo
    global blockcardcounter
    global atkcardcounter
    global strEX
    global bloEX
    blockcardcounter=0
    atkcardcounter=0
    str= PhotoImage(file="images/atk card.png")
    strEX = PhotoImage(file="images/atk card reskin.png")
    blo = PhotoImage(file="images/block card.png")
    bloEX = PhotoImage(file="images/atk card reskin.png")

    if ws2['A2'].value == "Strike":
        atkcardcounter=atkcardcounter+1

    if ws2['A2'].value == "Block":
        blockcardcounter = blockcardcounter+1

    if ws2['A3'].value == "Strike":
        atkcardcounter = atkcardcounter + 1


    if ws2['A3'].value == "Block":
        blockcardcounter = blockcardcounter + 1


    if ws2['A4'].value == "Strike":
        atkcardcounter = atkcardcounter + 1


    if ws2['A4'].value == "Block":
        blockcardcounter = blockcardcounter + 1


    if ws2['A5'].value == "Strike":
        atkcardcounter = atkcardcounter + 1


    if ws2['A5'].value == "Block":
        blockcardcounter = blockcardcounter + 1


    if ws2['A6'].value == "Strike":
        atkcardcounter = atkcardcounter + 1

    if ws2['A6'].value == "Block":
        blockcardcounter = blockcardcounter + 1

    if ws2['A7'].value == "Strike":
        atkcardcounter = atkcardcounter + 1

    if ws2['A7'].value == "Block":
        blockcardcounter = blockcardcounter + 1


    if ws2['A8'].value == "Strike":
        atkcardcounter = atkcardcounter + 1

    if ws2['A8'].value == "Block":
        blockcardcounter = blockcardcounter + 1

    if ws2['A9'].value == "Strike":
        atkcardcounter = atkcardcounter + 1

    if ws2['A9'].value == "Block":
        blockcardcounter = blockcardcounter + 1

    print("atk cards in Deck",atkcardcounter)
    print("Block cards in Deck",blockcardcounter)


    firstbscreen.after(900,auto_drawhand_animslot1)
    firstbscreen.after(1400, auto_drawhand_animslot2)
    firstbscreen.after(1900, auto_drawhand_animslot3)
    firstbscreen.after(2400, auto_drawhand_animslot4)
    firstbscreen.after(2900, auto_drawhand_animslot5)



def Endturnreset():
    global discardnumber
    global energynumber
    global Energylabel
    global discardlabel
    global slot1
    global slot2
    global slot3
    global slot4
    global slot5
    global slot1cardplayed
    global slot2cardplayed
    global slot3cardplayed
    global slot4cardplayed
    global slot5cardplayed
    global Shieldvar
    global shieldlabel
    global shieldnum
    global firstbscreen


    energynumber=3
    Energylabel.configure(text=(energynumber,"/3"))
# if shield have not been defined then the except will be played -----
    try:
        Shieldvar=0
        shieldlabel.configure(text=Shieldvar)
        shieldnum.place_forget()
        shieldlabel.place_forget()
    except:
        print("shields were not reset")

    try:
        if slot1cardplayed==True:
            slot1cardplayed=False
    except:
        print("slot 1 cleared")
        discardnumber = discardnumber + 1
        firstbscreen.after(100,slot1.destroy)

    else:
        slot1.place_forget()
    try:
        if slot2cardplayed==True:
            slot2cardplayed=False

    except:
        print("slot 2 cleared")
        firstbscreen.after(100, slot2.destroy)
        discardnumber=discardnumber+1

    else:
        slot2.place_forget()

    try:
        if slot3cardplayed==True:
            slot3cardplayed=False

    except:
        print("slot 3 cleared")
        firstbscreen.after(100, slot3.destroy)
        discardnumber=discardnumber+1

    else:
        slot3.place_forget()

    try:
        if slot4cardplayed==True:
            slot4cardplayed=False
    except:
        print("slot 4 cleared")
        firstbscreen.after(100, slot4.destroy)
        discardnumber=discardnumber+1

    else:
        slot4.place_forget()
    try:
        if slot5cardplayed==True:
            slot5cardplayed=False
    except:
        print("slot 5 cleared")
        firstbscreen.after(100, slot5.destroy)
        discardnumber=discardnumber+1

    else:
        slot5.place_forget()




    discardlabel.configure(text=discardnumber)
    firstbscreen.after(900,auto_drawhand_animslot1)
    firstbscreen.after(1400, auto_drawhand_animslot2)
    firstbscreen.after(1900, auto_drawhand_animslot3)
    firstbscreen.after(2400, auto_drawhand_animslot4)
    firstbscreen.after(2900, auto_drawhand_animslot5)



def update_character_anim2(ind):
    global framesE
    global frameCntE
    global enemy
    global firstbscreen

    frameE = framesE[ind]
    ind += 1

    if ind == frameCntE:
        ind = 0
    enemy.configure(image=frameE)
    firstbscreen.after(80, update_character_anim2, ind)

class ImageLabel(Label):
    """a label that displays images, and plays them if they are gifs"""
    def load(self, im):
        if isinstance(im, str):
            im = Image.open(im)
        self.loc = 0
        self.frames = []

        try:
            for i in count(1):
                self.frames.append(ImageTk.PhotoImage(im.copy()))
                im.seek(i)
        except EOFError:
            pass

        try:
            self.delay = im.info['duration']
        except:
            self.delay = 100

        if len(self.frames) == 1:
            self.config(image=self.frames[0])
        else:
            self.next_frame()

    def unload(self):
        self.config(image="")
        self.frames = None

    def next_frame(self):
        if self.frames:
            self.loc += 1
            self.loc %= len(self.frames)
            self.config(image=self.frames[self.loc])
            self.after(self.delay, self.next_frame)

def background_Track():
    pygame.mixer.music.stop()

def LanguageCheck():
    global wsL
    global wbL
    global Endturn
    if wsL['A1'].value =="Bengali":
        Endturn.config(text="এন্ড টার্ন")

    if wsL['A1'].value =="Japanese":
        Endturn.config(text="エンドターン")

    if wsL['A1'].value =="English":
        Endturn.config(text="End Turn")



def gamesetup():
    global firstbscreen
    global frames
    global frameCnt
    global frameCntE
    global framesE
    global mainguy
    global Drawlabel
    global Energylabel
    global discardlabel
    global energynumber
    global discardnumber
    global enemyhealth
    global enemyhealthlabel
    global enemyhealthvar
    global atksetup
    global health
    global healthlabel
    global healthvar
    global enemyatksetup
    global Shieldvar
    global cardclicked
    global shieldactivated
    global enemy
    global bg
    global enemyintentlabel
    global enemyhasdefence
    global Endturn
    bg = PhotoImage(file="images/background.png")
    dk = PhotoImage(file="images/deck.png")
    ds = PhotoImage(file="images/discard.png")
    en = PhotoImage(file="images/enemy.png")
    ener = PhotoImage(file="images/energy.png")
    mhb = PhotoImage(file="images/healthbar.png")
    ehb = PhotoImage(file="images/healthbar.png")
    qint = PhotoImage(file="images/unkown intent.PNG")
    coincolourbg = '#%02x%02x%02x' % (61,73,79)



    frameCntE = 12


    framesE = [PhotoImage(file='images/bird mob.gif', format='gif -index %i' % (i)) for i in range(frameCntE)]
    energynumber=3
    discardnumber=0
    enemyhealthvar=30
    healthvar=75
    Shieldvar=0
    cardclicked=0
    atksetup=True
    enemyatksetup=True
    shieldactivated=False
    enemyhasdefence=False

    # Show image using label
    label1 = Label(firstbscreen, image=bg)
    label1.place(x=0, y=0)

    mainguy=ImageLabel(firstbscreen)
    mainguy.place(relx=0.2, y=500, anchor="center")
    if wcS['A1'].value == "Defect":
        mainguy.load('images/The Defect.gif')
    if wcS['A1'].value == "Silent":
        mainguy.load("images/thief.gif")
    if wcS['A1'].value == "Ironclad":
        mainguy.load("images/Ironclad.gif")

    coinbag = PhotoImage(file="images/coinbag.png")
    coinbaglabel = Label(firstbscreen, image=coinbag)
    coinbaglabel.place(relx=0, y=0)
    coins = Label(firstbscreen, text="000", font=("Times New Roman", 30,), bg=coincolourbg, fg="yellow")
    coins.place(relx=0.05, y=2)

    enemy = Label(firstbscreen)
    enemy.place(relx=0.8, y=400, anchor="center")
    firstbscreen.after(0, update_character_anim2, 0)

    thedeck=Label(firstbscreen, image=dk)
    thedeck.place(relx=0.1, y=800, anchor="center")

    thediscard=Label(firstbscreen, image=ds)
    thediscard.place(relx=0.8, y=800, anchor="center")

    discardlabel = Label(firstbscreen, text=discardnumber, fg="blue",font=("Arial", 25))
    discardlabel.place(relx=0.8, y=800, anchor="center")

    Endturn = Button(firstbscreen,text="End Turn",bg="Dark Blue",fg="Grey",command=lambda:[Endturnreset(),enemydecision()])
    Endturn.place(relx=0.8, y=700, anchor="center")

    Drawlabel = Label(firstbscreen, text="X", fg="blue",font=("Arial", 25))
    Drawlabel.place(relx=0.1, y=800, anchor="center")

    Energy = Label(firstbscreen,image=ener)
    Energy.place(relx=0.07,y=650)

    Energylabel =Label(firstbscreen,text=(energynumber,"/3"),fg="green",bg="cyan",font=("Arial", 35))
    Energylabel.place(relx=0.075,y=670)

    health = Label(firstbscreen,image=mhb)
    health.place(relx=0.135,y=600)

    healthlabel = Label(firstbscreen,text=(healthvar,"/",healthvar),font=("Arial", 8))
    healthlabel.place(relx=0.126,y=600)

    enemyhealth = Label(firstbscreen,image=ehb)
    enemyhealth.place(relx=0.73,y=500)
    enemyhealthlabel = Label(firstbscreen, text=(enemyhealthvar,"/",enemyhealthvar), font=("Arial", 8))
    enemyhealthlabel.place(relx=0.72, y=500)

    enemyintentlabel = Label(firstbscreen, image=qint)
    enemyintentlabel.place(relx=0.805,y=240)
    background_Track()
    LanguageCheck()
    autodrawhand()
    firstbscreen.mainloop()



def firstbattlestart():
    global firstbscreen
    global ws2
    global wbL
    global wsL
    global wcS
    global wcL
    global pbL
    global psL
    wbL = load_workbook("Excel files/Settings.xlsx")
    wsL=wbL.active

    wb2 = load_workbook("Excel files/Active Deck.xlsx")
    ws2=wb2.active

    wcL = load_workbook("Excel files/Character Selection.xlsx")
    wcS = wcL.active

    firstbscreen=Toplevel()

    pbL = load_workbook("Excel files/Player Data.xlsx")
    psL=pbL.active

    firstbscreen.geometry("1500x875")
    gamesetup()
