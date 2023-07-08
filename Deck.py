from tkinter import *
from openpyxl import workbook, load_workbook

def lateruse(event):
    global button1
    global ws2
    global wb2
    global ds
    if ws2['E2'].value =="Delete":
        button1.grid_forget()
        ws2['E2'].value="normal"
        ws2['A2'].value="."
        wb2.save('Excel files/Active Deck.xlsx')
        ds.after(800,ds.destroy)

    else:
        print("not at a shop")

def lateruse2(event):
    global button2
    global ws2
    global wb2
    global ds
    if ws2['E2'].value =="Delete":
        button2.grid_forget()
        ws2['E2'].value="normal"
        ws2['A3'].value="."
        wb2.save('Excel files/Active Deck.xlsx')
        ds.after(800, ds.destroy)
    else:
        print("not at a shop")

def lateruse3(event):
    global button3
    global ws2
    global wb2
    global ds

    if ws2['E2'].value =="Delete":
        button3.grid_forget()
        ws2['E2'].value="normal"
        ws2['A4'].value="."
        wb2.save('Excel files/Active Deck.xlsx')
        ds.after(800, ds.destroy)
    else:
        print("not at a shop")

def lateruse4(event):
    global button4
    global ws2
    global wb2
    global ds

    if ws2['E2'].value =="Delete":
        button4.grid_forget()
        ws2['E2'].value="normal"
        ws2['A5'].value="."
        wb2.save('Excel files/Active Deck.xlsx')
        ds.after(800, ds.destroy)
    else:
        print("not at a shop")

def lateruse5(event):
    global button5
    global ws2
    global wb2
    global ds

    if ws2['E2'].value =="Delete":
        button5.grid_forget()
        ws2['E2'].value="normal"
        ws2['A6'].value="."
        wb2.save('Excel files/Active Deck.xlsx')
        ds.after(800, ds.destroy)
    else:
        print("not at a shop")

def lateruse6(event):
    global button6
    global ws2
    global wb2
    global ds

    if ws2['E2'].value =="Delete":
        button6.grid_forget()
        ws2['E2'].value="normal"
        ws2['A7'].value="."
        wb2.save('Excel files/Active Deck.xlsx')
        ds.after(800, ds.destroy)
    else:
        print("not at a shop")

def lateruse7(event):
    global button7
    global ws2
    global wb2
    global ds

    if ws2['E2'].value =="Delete":
        button7.grid_forget()
        ws2['E2'].value="normal"
        ws2['A8'].value="."
        wb2.save('Excel files/Active Deck.xlsx')
        ds.after(800, ds.destroy)
    else:
        print("not at a shop")

def lateruse8(event):
    global button8
    global ws2
    global wb2
    global ds

    if ws2['E2'].value =="Delete":
        button8.grid_forget()
        ws2['E2'].value="normal"
        ws2['A9'].value="."
        wb2.save('Excel files/Active Deck.xlsx')
        ds.after(800, ds.destroy)
    else:
        print("not at a shop")

def lateruse9(event):
    global button9
    global ws2
    global wb2
    global ds

    if ws2['E2'].value =="Delete":
        button9.grid_forget()
        ws2['E2'].value="normal"
        ws2['A10'].value="."
        wb2.save('Excel files/Active Deck.xlsx')
        ds.after(800, ds.destroy)
    else:
        print("not at a shop")


def lateruse10(event):
    global button10
    global ws2
    global wb2
    global ds

    if ws2['E2'].value =="Delete":
        button10.grid_forget()
        ws2['E2'].value="normal"
        ws2['A11'].value="."
        wb2.save('Excel files/Active Deck.xlsx')
        ds.after(800, ds.destroy)
    else:
        print("not at a shop")

def descript(event):
    global ws
    global description

    description.config(text=ws['B2'].value)

def descript2(event):
    global ws
    global description

    description.config(text=ws['B3'].value)

def descript3(event):
    global ws
    global description

    description.config(text=ws['B4'].value)

def descript4(event):
    global ws
    global description

    description.config(text=ws['B5'].value)

def descript5(event):
    global ws
    global description

    description.config(text=ws['B6'].value)

def descript6(event):
    global ws
    global description

    description.config(text=ws['B7'].value)

def descript7(event):
    global ws
    global description

    description.config(text=ws['B8'].value)

def descript8(event):
    global ws
    global description

    description.config(text=ws['B9'].value)


def collection():
    global ds
    global mainarea
    global sidebar
    global description
    global ws2
    global button1
    global button2
    global button3
    global button4
    global button5
    global button6
    global button7
    global button8
    global button9
    global button10
    card1 = PhotoImage(file="images/atk card.png")
    card2 = PhotoImage(file="images/block card.png")
    bigblock=PhotoImage(file="images/30 block card.png")
    repair=PhotoImage(file="images/Self repair card.png")
    heavyhit=PhotoImage(file="images/Heavy hit card reskin.png")
    multihit=PhotoImage(file="images/multi slash card.png")
    doublestr=PhotoImage(file="images/Double Strength card.png")
    more_energy=PhotoImage(file="images/Double Energy.png")

    if ws2['A2'].value =="Strike":
        button1 = Button(mainarea, image=card1)
        button1.bind('<Button-1>', descript)
        button1.bind('<Double-1>', lateruse)
        button1.grid(row=1, column=1)

    if ws2['A2'].value =="Block":
        button1 = Button(mainarea, image=card2)
        button1.bind('<Button-1>', descript2)
        button1.bind('<Double-1>', lateruse)
        button1.grid(row=1, column=1)

    if ws2['A3'].value =="Strike":
        button2 = Button(mainarea, image=card1)
        button2.bind('<Button-1>', descript)
        button2.bind('<Double-1>', lateruse2)
        button2.grid(row=1, column=2)

    if ws2['A3'].value =="Block":
        button2 = Button(mainarea, image=card2)
        button2.bind('<Button-1>', descript2)
        button2.bind('<Double-1>', lateruse2)
        button2.grid(row=1, column=2)

    if ws2['A4'].value =="Strike":
        button3 = Button(mainarea, image=card1)
        button3.bind('<Button-1>', descript)
        button3.bind('<Double-1>', lateruse3)
        button3.grid(row=1, column=3)

    if ws2['A4'].value =="Block":
        button3 = Button(mainarea, image=card2)
        button3.bind('<Button-1>', descript2)
        button3.bind('<Double-1>', lateruse3)
        button3.grid(row=1, column=3)

    if ws2['A5'].value =="Strike":
        button4 = Button(mainarea, image=card1)
        button4.bind('<Button-1>', descript)
        button4.bind('<Double-1>', lateruse4)
        button4.grid(row=1, column=4)

    if ws2['A5'].value =="Block":
        button4 = Button(mainarea, image=card2)
        button4.bind('<Button-1>', descript2)
        button4.bind('<Double-1>', lateruse4)
        button4.grid(row=1, column=4)

    if ws2['A6'].value == "Strike":
        button5 = Button(mainarea, image=card1)
        button5.bind('<Button-1>', descript)
        button5.bind('<Double-1>', lateruse5)
        button5.grid(row=2, column=1)

    if ws2['A6'].value == "Block":
        button5 = Button(mainarea, image=card2)
        button5.bind('<Button-1>', descript2)
        button5.bind('<Double-1>', lateruse5)
        button5.grid(row=2, column=1)

    if ws2['A7'].value == "Strike":
        button6 = Button(mainarea, image=card1)
        button6.bind('<Button-1>', descript)
        button6.bind('<Double-1>', lateruse6)
        button6.grid(row=2, column=2)

    if ws2['A7'].value == "Block":
        button6 = Button(mainarea, image=card2)
        button6.bind('<Button-1>', descript2)
        button6.bind('<Double-1>', lateruse6)
        button6.grid(row=2, column=2)

    if ws2['A8'].value == "Strike":
        button7 = Button(mainarea, image=card1)
        button7.bind('<Button-1>', descript)
        button7.bind('<Double-1>', lateruse7)
        button7.grid(row=2, column=3)

    if ws2['A8'].value == "Block":
        button7 = Button(mainarea, image=card2)
        button7.bind('<Button-1>', descript2)
        button7.bind('<Double-1>', lateruse7)
        button7.grid(row=2, column=3)

    if ws2['A9'].value == "Strike":
        button8 = Button(mainarea, image=card1)
        button8.bind('<Button-1>', descript)
        button8.bind('<Double-1>', lateruse8)
        button8.grid(row=2, column=4)

    if ws2['A9'].value == "Block":
        button8 = Button(mainarea, image=card2)
        button8.bind('<Button-1>', descript2)
        button8.bind('<Double-1>', lateruse8)
        button8.grid(row=2, column=4)

    if ws2['A10'].value == "Double":
        button9 = Button(mainarea, image=doublestr)
        button9.bind('<Button-1>', descript4)
        button9.bind('<Double-1>', lateruse9)
        button9.grid(row=3, column=1)

    if ws2['A10'].value == "30block":
        button9 = Button(mainarea, image=bigblock)
        button9.bind('<Button-1>', descript3)
        button9.bind('<Double-1>', lateruse9)
        button9.grid(row=3, column=1)

    if ws2['A10'].value == "multi":
        button9 = Button(mainarea, image=multihit)
        button9.bind('<Button-1>', descript5)
        button9.bind('<Double-1>', lateruse9)
        button9.grid(row=3, column=1)

    if ws2['A10'].value == "repair":
        button9 = Button(mainarea, image=repair)
        button9.bind('<Button-1>', descript6)
        button9.bind('<Double-1>', lateruse9)
        button9.grid(row=3, column=1)

    if ws2['A10'].value == "heavy":
        button9 = Button(mainarea, image=heavyhit)
        button9.bind('<Button-1>', descript7)
        button9.bind('<Double-1>', lateruse9)
        button9.grid(row=3, column=1)

    if ws2['A10'].value == "moreE":
        button9 = Button(mainarea, image=more_energy)
        button9.bind('<Button-1>', descript8)
        button9.bind('<Double-1>', lateruse9)
        button9.grid(row=3, column=1)

    if ws2['A11'].value == "Double":
        button10 = Button(mainarea, image=doublestr)
        button10.bind('<Button-1>', descript4)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=3, column=2)

    if ws2['A11'].value == "30block":
        button10 = Button(mainarea, image=bigblock)
        button10.bind('<Button-1>', descript3)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=3, column=2)

    if ws2['A11'].value == "multi":
        button10 = Button(mainarea, image=multihit)
        button10.bind('<Button-1>', descript5)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=3, column=2)

    if ws2['A11'].value == "repair":
        button10 = Button(mainarea, image=repair)
        button10.bind('<Button-1>', descript6)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=3, column=2)

    if ws2['A11'].value == "heavy":
        button10 = Button(mainarea, image=heavyhit)
        button10.bind('<Button-1>', descript7)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=3, column=2)

    if ws2['A11'].value == "moreE":
        button10 = Button(mainarea, image=more_energy)
        button10.bind('<Button-1>', descript8)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=3, column=2)

    if ws2['A12'].value == "Double":
        button10 = Button(mainarea, image=doublestr)
        button10.bind('<Button-1>', descript4)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=3, column=3)

    if ws2['A12'].value == "30block":
        button10 = Button(mainarea, image=bigblock)
        button10.bind('<Button-1>', descript3)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=3, column=3)

    if ws2['A12'].value == "multi":
        button10 = Button(mainarea, image=multihit)
        button10.bind('<Button-1>', descript5)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=3, column=3)

    if ws2['A12'].value == "repair":
        button10 = Button(mainarea, image=repair)
        button10.bind('<Button-1>', descript6)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=3, column=3)

    if ws2['A12'].value == "heavy":
        button10 = Button(mainarea, image=heavyhit)
        button10.bind('<Button-1>', descript7)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=3, column=3)

    if ws2['A12'].value == "moreE":
        button10 = Button(mainarea, image=more_energy)
        button10.bind('<Button-1>', descript8)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=3, column=3)

    if ws2['A13'].value == "Double":
        button10 = Button(mainarea, image=doublestr)
        button10.bind('<Button-1>', descript4)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=3, column=4)

    if ws2['A13'].value == "30block":
        button10 = Button(mainarea, image=bigblock)
        button10.bind('<Button-1>', descript3)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=3, column=4)

    if ws2['A13'].value == "multi":
        button10 = Button(mainarea, image=multihit)
        button10.bind('<Button-1>', descript5)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=3, column=4)

    if ws2['A13'].value == "repair":
        button10 = Button(mainarea, image=repair)
        button10.bind('<Button-1>', descript6)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=3, column=4)

    if ws2['A13'].value == "heavy":
        button10 = Button(mainarea, image=heavyhit)
        button10.bind('<Button-1>', descript7)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=3, column=4)

    if ws2['A13'].value == "moreE":
        button10 = Button(mainarea, image=more_energy)
        button10.bind('<Button-1>', descript8)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=3, column=4)

    if ws2['A14'].value == "Double":
        button10 = Button(mainarea, image=doublestr)
        button10.bind('<Button-1>', descript4)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=4, column=1)

    if ws2['A14'].value == "30block":
        button10 = Button(mainarea, image=bigblock)
        button10.bind('<Button-1>', descript3)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=4, column=1)

    if ws2['A14'].value == "multi":
        button10 = Button(mainarea, image=multihit)
        button10.bind('<Button-1>', descript5)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=4, column=1)

    if ws2['A14'].value == "repair":
        button10 = Button(mainarea, image=repair)
        button10.bind('<Button-1>', descript6)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=4, column=1)

    if ws2['A14'].value == "heavy":
        button10 = Button(mainarea, image=heavyhit)
        button10.bind('<Button-1>', descript7)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=4, column=1)

    if ws2['A14'].value == "moreE":
        button10 = Button(mainarea, image=more_energy)
        button10.bind('<Button-1>', descript8)
        button10.bind('<Double-1>', lateruse10)
        button10.grid(row=4, column=1)

    abovedesc=Label(sidebar,text="Description:")
    abovedesc.pack()
    description=Label(sidebar,text="")
    description.pack()

    ds.mainloop()


def Deckstart():
    global ds
    global mainarea
    global sidebar
    global ws
    global ws2
    global wb2
    global firstloadvar

    wb = load_workbook("Excel files/Cards.xlsx")
    wb2 = load_workbook("Excel files/Active Deck.xlsx")
    ws = wb.active
    ws2 = wb2.active
    ds=Toplevel()
    if ws2['E2'].value =="Delete":
        ds.title("CHOOSE A CARD TO REMOVE FROM YOUR DECK PERMANENTLY")
    else:
        ds.title("The Deck")
    mainarea = Frame(ds, bg='#CCC', width=500, height=500)
    mainarea.pack(expand=True, fill='both', side='left')

    if ws2['E2'].value =="Delete":
        sidebar = Frame(ds, width=200, bg='Black', height=300, relief='sunken', borderwidth=2)
        sidebar.pack(expand=False, fill='both', side='right', anchor='nw')

    else:
        sidebar = Frame(ds, width=200, bg='Dark grey', height=300, relief='sunken', borderwidth=2)
        sidebar.pack(expand=False, fill='both', side='right', anchor='nw')


    collection()
