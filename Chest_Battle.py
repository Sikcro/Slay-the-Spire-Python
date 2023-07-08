from tkinter import *
from openpyxl import workbook, load_workbook
import random
import pygame
from PIL import Image, ImageTk
from itertools import count
from RewardScreenlvl2 import Rewardstartlvl2
from Death import PLayerDied

def gameover():
    global firstbscreen
    firstbscreen.destroy()

    PLayerDied()


def RewardScreen():
    global firstbscreen
    global thespearmanhasfallen
    global thebandithasfallen
    from TheWorld import THEboard
    if thebandithasfallen and thespearmanhasfallen ==True:
        firstbscreen.destroy()
        THEboard()

    if thespearmanhasfallen ==True:
        print("spear guy dead, one enemy left")

    if thebandithasfallen ==True:
        print("bandit guy dead, one enemy left")

def secondenemydefend():
    global enemy2hasdefence
    global secondenemyshieldvar
    global secondenemyshieldbar


    enemy2hasdefence=True
    shieldbarpng=PhotoImage(file="images/shield.png")


    secondenemyshieldbar=Label(firstbscreen,image=shieldbarpng)
    secondenemyshieldbar.place(relx=0.53,y=570)
    Eshield=6

    secondenemyshieldvar=Label(firstbscreen,text=Eshield)
    secondenemyshieldvar.place(relx=0.52,y=570)
    firstbscreen.mainloop()


def enemydefend():
    global enemyhasdefence
    global enemyshieldbar
    global enemyshieldvar


    enemyhasdefence=True
    shieldbarpng=PhotoImage(file="images/shield.png")
    enemyshieldbar=Label(firstbscreen,image=shieldbarpng)
    enemyshieldbar.place(relx=0.73,y=530)

    Eshield=6
    enemyshieldvar=Label(firstbscreen,text=Eshield)
    enemyshieldvar.place(relx=0.72,y=530)


    firstbscreen.mainloop()



def enemyatk():
    global health
    global healthlabel
    global healthvar
    global Efirstatkcomplete
    global Esecondatkcomplete
    global Ethirdatkcomplete
    global Efourthatkcomplete
    global Efithatkcomplete
    global Esixthatkcomplete
    global Eeighthatkcomplete
    global Eseventhatkcomplete
    global enemyatksetup
    global shieldactivated
    global Shieldvar
    global shieldlabel
    global shieldnum
    seventypercent=PhotoImage(file="images/healthbar 70 percent dmg.png")
    ninetypercent=PhotoImage(file="images/healthbar 90 percent dmg.png")
    fiftypercent=PhotoImage(file="images/healthbar half dmg.png")
    twentypercent=PhotoImage(file="images/healthbar small dmg.png")
    onehundredpercent=PhotoImage(file="images/healthbar 100 percent dmg.png")
    enemyatkpower=10

    try:
        if shieldactivated == True:
            enemyatkpower=0
            Shieldvar=0
            shieldnum.destroy()
            shieldlabel.destroy()

        if enemyatkpower<0:
            enemyatkpower=0
            print("working")
        else:
            print("enemy pwr not less than 0")
        healthvar=healthvar-enemyatkpower

    except:
        print("no shield played this turn")




    if healthvar <=0:
        gameover()

    else:
        try:
            if Eseventhatkcomplete == True:

                if shieldactivated == True:
                    healthlabel.configure(text=(healthvar, "/75"))
                    shieldactivated=False

                if healthvar > 0:
                    healthlabel.configure(text=(healthvar, "/75"))

                else:
                    health.configure(image=onehundredpercent)
                    healthlabel.configure(text=(healthvar,"/75"))
                    Eseventhatkcomplete = False
                    print("YOU lost")
                    gameover()

        except:
            print(".")




        try:
            if Esixthatkcomplete == True:
                if shieldactivated == True:
                    healthlabel.configure(text=(healthvar, "/75"))
                    shieldactivated=False

                else:
                    health.configure(image=onehundredpercent)
                    healthlabel.configure(text=(healthvar,"/75"))
                    Esixthatkcomplete = False
                    Eseventhatkcomplete = True
        except:
            print(".")



        try:
            if Efithatkcomplete == True:
                if shieldactivated == True:
                    healthlabel.configure(text=(healthvar, "/75"))
                    shieldactivated=False

                else:
                    health.configure(image=ninetypercent)
                    healthlabel.configure(text=(healthvar,"/75"))
                    Efithatkcomplete = False
                    Esixthatkcomplete= True

        except:
            print(".")


        try:
            if Efourthatkcomplete==True:

                if shieldactivated == True:
                    healthlabel.configure(text=(healthvar, "/75"))
                    shieldactivated=False
                else:
                    health.configure(image=ninetypercent)
                    healthlabel.configure(text=(healthvar,"/75"))
                    Efourthatkcomplete=False
                    Efithatkcomplete=True

        except:
            print(".")



        try:
            if Ethirdatkcomplete==True:

                if shieldactivated == True:
                    healthlabel.configure(text=(healthvar, "/75"))
                    shieldactivated=False
                else:
                    health.configure(image=ninetypercent)
                    healthlabel.configure(text=(healthvar,"/75"))
                    Ethirdatkcomplete=False
                    Efourthatkcomplete=True
        except:
            print(".")


        try:
            if Esecondatkcomplete==True:
                if shieldactivated == True:
                    healthlabel.configure(text=(healthvar, "/75"))
                    shieldactivated=False
                else:
                    health.configure(image=seventypercent)
                    healthlabel.configure(text=(healthvar,"/75"))
                    Ethirdatkcomplete=True
        except:
            print(".")

        try:
            if Efirstatkcomplete==True:

                if shieldactivated == True:
                    healthlabel.configure(text=(healthvar, "/75"))
                    shieldactivated=False
                else:
                    health.configure(image=fiftypercent)
                    healthlabel.configure(text=(healthvar,"/75"))
                    Efirstatkcomplete=False
                    Esecondatkcomplete=True
        except:
            print(".")

        if  enemyatksetup==True:

            if shieldactivated==True:
                healthlabel.configure(text=(healthvar, "/75"))
                shieldactivated=False

            else:
                health.configure(image=twentypercent)
                healthlabel.configure(text=(healthvar,"/75"))
                Efirstatkcomplete=True
                enemyatksetup=False

        else:
            print(".")

    firstbscreen.mainloop()


def enemy_decision2():
    global secondenemyintent
    global firstbscreen
    global thebandithasfallen

    if thebandithasfallen==True:
        enemy2dice = 0
        print("the spear dice value is", enemy2dice)

    else:
        enemy2dice = random.randint(1, 2)

    print("enemy2dice result is ", enemy2dice)
    atkintent = PhotoImage(file="images/atk intent.png")
    defintent = PhotoImage(file="images/shield intent.PNG")
    deadintent = PhotoImage(file="images/Dead mob.png")
    if enemy2dice == 1:
        secondenemyintent.config(image=atkintent)

        enemyatk()

    if enemy2dice == 2:
        secondenemyintent.config(image=defintent)
        secondenemydefend()

    if enemy2dice ==0:
        secondenemyintent.config(image=deadintent)
        print("This mob is Dead")

    firstbscreen.mainloop()
def enemydecision():
    global enemyintentlabel
    global firstbscreen
    global thespearmanhasfallen

    if thespearmanhasfallen==True:
        enemydice = 0
        print("the spear dice value is",enemydice)

    else:
        enemydice=random.randint(1,2)
    atkintent = PhotoImage(file="images/atk intent.png")
    defintent = PhotoImage(file="images/shield intent.PNG")
    deadintent = PhotoImage(file="images/Dead mob.png")
    if enemydice ==1:
        enemyintentlabel.config(image=atkintent)

        enemyatk()
        firstbscreen.after(150,enemy_decision2)

    if enemydice ==2:
        enemyintentlabel.config(image=defintent)
        firstbscreen.after(150, enemy_decision2)

        enemydefend()

    if enemydice==0:
        enemyintentlabel.config(image=deadintent)
        print("This mob is Dead")

        firstbscreen.after(150, enemy_decision2)


    firstbscreen.mainloop()




def onetimeuse():
    global wb2
    global ws2
    thercard = ws2['A10']
    thercard.value = "used"
    wb2.save('Excel files/Active Deck.xlsx')
    print("one time used up")

def repaircard():
    global healthvar
    global firstbscreen
    healthvar=healthvar+7
    healthlabel.configure(text=(healthvar, "/75"))
    print("working...")
    firstbscreen.after(200,onetimeuse)
    firstbscreen.mainloop()

def multihitcard():
    global firstbscreen
    firstbscreen.after(100,strikecardforsecondenemy)
    firstbscreen.after(200,strikecard)


def bigblockcard():
    global firstbscreen
    global Shieldvar
    global shieldactivated
    global shieldlabel
    global shieldnum
    global cardclicked
    shieldactivated=True
    shieldimg=PhotoImage(file="images/shield.png")
    cardclicked=cardclicked+1

    if energynumber <0:
        print("cannot apply Block")

    else:

        if cardclicked==1:
            try:
                shieldlabel.place_forget()
                shieldnum.place_forget()
            except:
                print("-")
            shieldlabel=Label(firstbscreen,image=shieldimg)
            shieldlabel.place(relx=0.135, y=630)
            Shieldvar=Shieldvar+30
            shieldnum=Label(firstbscreen,text=Shieldvar)
            shieldnum.place(relx=0.135, y=630)
        if cardclicked==2:
            try:
                shieldlabel.place_forget()
                shieldnum.place_forget()
            except:
                print("-")
            shieldlabel=Label(firstbscreen,image=shieldimg)
            shieldlabel.place(relx=0.135, y=630)
            Shieldvar=Shieldvar+30
            shieldnum=Label(firstbscreen,text=Shieldvar)
            shieldnum.place(relx=0.135, y=630)

        if cardclicked==3:
            try:
                shieldlabel.place_forget()
                shieldnum.place_forget()
            except:
                print("-")

            shieldlabel=Label(firstbscreen,image=shieldimg)
            shieldlabel.place(relx=0.135, y=630)
            Shieldvar=Shieldvar+30
            shieldnum=Label(firstbscreen,text=Shieldvar)
            shieldnum.place(relx=0.135, y=630)
            cardclicked=0

    firstbscreen.mainloop()



def heavycard_secondenenmy():
    global enemy2healthlabel
    global enemy2healthvar
    global secondenemyshieldbar
    global secondenemyshieldvar
    global firstbscreen
    global secenemy_atksetup
    global secenemy_firstatkcomplete
    global secenemy_secondatkcomplete
    global secenemy_thirdatkcomplete
    global secenemy_fourthatkcomplete
    global bg
    global enemy2hasdefence
    global energynumber
    global doublestrength
    global enemyshieldvar
    global enemyshieldbar
    global thebandithasfallen
    global banditmanchoice
    global spearmanchoice
    global secondenemyintent
    try:
        banditmanchoice.place_forget()
        spearmanchoice.place_forget()
    except:
        print("-------------------")
    seventypercent=PhotoImage(file="images/healthbar 70 percent dmg.png")
    ninetypercent=PhotoImage(file="images/healthbar 90 percent dmg.png")
    fiftypercent=PhotoImage(file="images/healthbar half dmg.png")
    twentypercent=PhotoImage(file="images/healthbar small dmg.png")
    onehundredpercent=PhotoImage(file="images/healthbar 100 percent dmg.png")
    reward=PhotoImage(file="images/win screen.png")
    attackpower=12

    if energynumber <0:
        print("cannot attack")

    else:
        enemy2healthvar=enemy2healthvar-attackpower

        if enemy2healthvar <= 0:
            thebandithasfallen=True
            Deadintent = PhotoImage(file="images/Dead mob.png")
            secondenemyintent.config(image=Deadintent)
            RewardScreen()

        try:
            if enemy2hasdefence == True:
                try:
                    secondenemyshieldvar.place_forget()
                    secondenemyshieldbar.place_forget()

                except:
                    print("bandit shield not spawned")


        except:
            print("enemy has no shield")




        try:
            if secenemy_thirdatkcomplete==True:

                if enemy2hasdefence==True:
                    enemy2healthvar = enemy2healthvar + attackpower
                    enemy2health.configure(image=seventypercent)
                    enemy2healthlabel.configure(text=enemy2healthvar)
                    secenemy_thirdatkcomplete=True
                    enemy2hasdefence=False
                else:
                    Deadintent = PhotoImage(file="images/Dead mob.png")
                    enemy2health.configure(image=onehundredpercent)
                    enemy2healthlabel.configure(text=enemy2healthvar)
                    secenemy_thirdatkcomplete=False
                    thebandithasfallen = True
                    secondenemyintent.config(image=Deadintent)
                    RewardScreen()



        except:
            print(".")


        try:
            if secenemy_secondatkcomplete==True:

                if enemy2hasdefence==True:
                    enemy2healthvar = enemyhealthvar + attackpower
                    enemy2health.configure(image=fiftypercent)
                    enemy2healthlabel.configure(text=enemy2healthvar)
                    secenemy_secondatkcomplete=True
                    secenemy_thirdatkcomplete=False
                    enemy2hasdefence=False
                else:
                    enemy2health.configure(image=seventypercent)
                    enemy2healthlabel.configure(text=enemy2healthvar)
                    secenemy_secondatkcomplete=False
                    secenemy_thirdatkcomplete=True

        except:
            print(".")

        try:
            if secenemy_firstatkcomplete==True:

                if enemy2hasdefence==True:
                    enemy2healthvar=enemy2healthvar+attackpower
                    enemy2health.configure(image=twentypercent)
                    enemy2healthlabel.configure(text=enemy2healthvar)
                    secenemy_secondatkcomplete=False
                    secenemy_firstatkcomplete=True
                    enemy2hasdefence=False
                else:
                    enemy2health.configure(image=fiftypercent)
                    enemy2healthlabel.configure(text=enemy2healthvar)
                    secenemy_firstatkcomplete=False
                    secenemy_secondatkcomplete=True

        except:
            print(".")

        if  secenemy_atksetup==True:
            enemy2health.configure(image=twentypercent)
            enemy2healthlabel.configure(text=enemy2healthvar)
            secenemy_firstatkcomplete=True
            secenemy_atksetup=False


        else:
            print(".")


    firstbscreen.mainloop()


def heavycard():
    global enemyhealthlabel
    global enemyhealth
    global enemyhealthvar
    global firstbscreen
    global atksetup
    global firstatkcomplete
    global secondatkcomplete
    global thirdatkcomplete
    global fourthatkcomplete
    global bg
    global enemyhasdefence
    global enemyshieldvar
    global enemyshieldbar
    global secondenemyshieldbar
    global secondenemyshieldvar
    global energynumber

    seventypercent=PhotoImage(file="images/healthbar 70 percent dmg.png")
    ninetypercent=PhotoImage(file="images/healthbar 90 percent dmg.png")
    fiftypercent=PhotoImage(file="images/healthbar half dmg.png")
    twentypercent=PhotoImage(file="images/healthbar small dmg.png")
    onehundredpercent=PhotoImage(file="images/healthbar 100 percent dmg.png")
    reward=PhotoImage(file="images/win screen.png")
    attackpower=12

    try:
        spearmanchoice.place_forget()
        banditmanchoice.place_forget()
    except:
        print("--------------------------")


    if energynumber <0:
        print("cannot attack")

    else:

        enemyhealthvar=enemyhealthvar-attackpower
        if enemyhealthvar <= 0:
            print("YOU won!!!")
            RewardScreen()

        try:
            if enemyhasdefence == True:
                try:
                    enemyshieldvar.place_forget()
                    enemyshieldbar.place_forget()

                except:
                    print("bandit shield not spawned")
            if enemy2hasdefence == True:
                try:
                    secondenemyshieldvar.place_forget()
                    secondenemyshieldbar.place_forget()

                except:
                    print("bandit shield not spawned")


        except:
            print("enemy has no shield")



        try:
            if fourthatkcomplete==True:

                if enemyhasdefence==True:
                    enemyhealthvar = enemyhealthvar + attackpower
                    enemyhealth.configure(image=ninetypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    fourthatkcomplete = True
                    enemyhasdefence=False

                else:
                    enemyhealth.configure(image=onehundredpercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)

                    fourthatkcomplete=False
                    print("YOU won!!!")
                    RewardScreen()




        except:
            print(".")



        try:
            if thirdatkcomplete==True:

                if enemyhasdefence==True:
                    enemyhealthvar = enemyhealthvar + attackpower
                    enemyhealth.configure(image=seventypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    thirdatkcomplete=True
                    fourthatkcomplete=False
                    enemyhasdefence=False
                else:
                    enemyhealth.configure(image=ninetypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    thirdatkcomplete=False
                    fourthatkcomplete=True


        except:
            print(".")


        try:
            if secondatkcomplete==True:

                if enemyhasdefence==True:
                    enemyhealthvar = enemyhealthvar + attackpower
                    enemyhealth.configure(image=fiftypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    secondatkcomplete=True
                    thirdatkcomplete=False
                    enemyhasdefence=False
                else:
                    enemyhealth.configure(image=seventypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    secondatkcomplete=False
                    thirdatkcomplete=True

        except:
            print(".")

        try:
            if firstatkcomplete==True:

                if enemyhasdefence==True:
                    enemyhealthvar=enemyhealthvar+attackpower
                    enemyhealth.configure(image=twentypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    secondatkcomplete=False
                    firstatkcomplete=True
                    enemyhasdefence=False
                else:
                    enemyhealth.configure(image=fiftypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    firstatkcomplete=False
                    secondatkcomplete=True

        except:
            print(".")

        if  atksetup==True:
            enemyhealth.configure(image=twentypercent)
            enemyhealthlabel.configure(text=enemyhealthvar)
            firstatkcomplete=True
            atksetup=False


        else:
            print(".")


    firstbscreen.mainloop()

def doublestrcard():
    global doublestrength
    doublestrength=True

def moreEcard():
    global energynumber
    global EnergyLabel
    global firstbscreen
    energynumber=energynumber+3
    Energylabel.configure(text=(energynumber, "/3"))
    onetimeuse()
    firstbscreen.mainloop()

def defencecard():
    global firstbscreen
    global Shieldvar
    global shieldactivated
    global shieldlabel
    global shieldnum
    global cardclicked
    shieldactivated=True
    shieldimg=PhotoImage(file="images/shield.png")
    cardclicked=cardclicked+1

    if energynumber <0:
        print("cannot apply Block")

    else:
        pygame.mixer.init()
        pygame.mixer.music.load("Audio/r_laser_p-95552.mp3")
        pygame.mixer.music.play(loops=1)

        if cardclicked==1:
            try:
                shieldlabel.place_forget()
                shieldnum.place_forget()
            except:
                print("-")
            shieldlabel=Label(firstbscreen,image=shieldimg)
            shieldlabel.place(relx=0.135, y=630)
            Shieldvar=Shieldvar+5
            shieldnum=Label(firstbscreen,text=Shieldvar)
            shieldnum.place(relx=0.135, y=630)
        if cardclicked==2:
            try:
                shieldlabel.place_forget()
                shieldnum.place_forget()
            except:
                print("-")
            shieldlabel=Label(firstbscreen,image=shieldimg)
            shieldlabel.place(relx=0.135, y=630)
            Shieldvar=Shieldvar+5
            shieldnum=Label(firstbscreen,text=Shieldvar)
            shieldnum.place(relx=0.135, y=630)

        if cardclicked==3:
            try:
                shieldlabel.place_forget()
                shieldnum.place_forget()
            except:
                print("-")

            shieldlabel=Label(firstbscreen,image=shieldimg)
            shieldlabel.place(relx=0.135, y=630)
            Shieldvar=Shieldvar+5
            shieldnum=Label(firstbscreen,text=Shieldvar)
            shieldnum.place(relx=0.135, y=630)
            cardclicked=0

    firstbscreen.mainloop()


def strikecardforsecondenemy():
    global enemy2healthlabel
    global enemy2healthvar
    global secondenemyshieldbar
    global secondenemyshieldvar
    global firstbscreen
    global secenemy_atksetup
    global secenemy_firstatkcomplete
    global secenemy_secondatkcomplete
    global secenemy_thirdatkcomplete
    global secenemy_fourthatkcomplete
    global bg
    global enemyhasdefence
    global enemy2hasdefence
    global energynumber
    global doublestrength
    global enemyshieldvar
    global enemyshieldbar
    global thebandithasfallen
    global banditmanchoice
    global spearmanchoice
    global secondenemyintent
    global selclick
    global enemy2
    try:
        banditmanchoice.place_forget()
        spearmanchoice.place_forget()
        selclick=0
    except:
        print("-------------------")
    seventypercent=PhotoImage(file="images/healthbar 70 percent dmg.png")
    ninetypercent=PhotoImage(file="images/healthbar 90 percent dmg.png")
    fiftypercent=PhotoImage(file="images/healthbar half dmg.png")
    twentypercent=PhotoImage(file="images/healthbar small dmg.png")
    onehundredpercent=PhotoImage(file="images/healthbar 100 percent dmg.png")
    reward=PhotoImage(file="images/win screen.png")
    if doublestrength==True:
        attackpower=6*2
        doublestrength=False

    else:
        attackpower=6

    if energynumber <0:
        print("cannot attack")

    else:
        pygame.mixer.init()
        pygame.mixer.music.load("Audio/cartoon punch 1.mp3")
        pygame.mixer.music.play(loops=1)
        enemy2healthvar=enemy2healthvar-attackpower

        if enemy2healthvar <= 0:
            thebandithasfallen=True
            Deadintent = PhotoImage(file="images/Dead mob.png")
            secondenemyintent.config(image=Deadintent)
            enemy2.place_forget()
            enemy2healthlabel.place_forget()
            enemy2health.place_forget()
            RewardScreen()

        try:
            if enemy2hasdefence == True:
                try:
                    secondenemyshieldvar.place_forget()
                    secondenemyshieldbar.place_forget()

                except:
                    print("bandit shield not spawned")


        except:
            print("enemy has no shield")




        try:
            if secenemy_thirdatkcomplete==True:

                if enemy2hasdefence==True:
                    enemy2healthvar = enemy2healthvar + attackpower
                    enemy2health.configure(image=seventypercent)
                    enemy2healthlabel.configure(text=enemy2healthvar)
                    secenemy_thirdatkcomplete=True
                    enemy2hasdefence=False
                else:
                    Deadintent = PhotoImage(file="images/Dead mob.png")
                    enemy2health.configure(image=onehundredpercent)
                    enemy2healthlabel.configure(text=enemy2healthvar)
                    secenemy_thirdatkcomplete=False
                    thebandithasfallen = True
                    secondenemyintent.config(image=Deadintent)
                    RewardScreen()



        except:
            print(".")


        try:
            if secenemy_secondatkcomplete==True:

                if enemy2hasdefence==True:
                    enemy2healthvar = enemyhealthvar + attackpower
                    enemy2health.configure(image=fiftypercent)
                    enemy2healthlabel.configure(text=enemy2healthvar)
                    secenemy_secondatkcomplete=True
                    secenemy_thirdatkcomplete=False
                    enemy2hasdefence=False
                else:
                    enemy2health.configure(image=seventypercent)
                    enemy2healthlabel.configure(text=enemy2healthvar)
                    secenemy_secondatkcomplete=False
                    secenemy_thirdatkcomplete=True

        except:
            print(".")

        try:
            if secenemy_firstatkcomplete==True:

                if enemy2hasdefence==True:
                    enemy2healthvar=enemy2healthvar+attackpower
                    enemy2health.configure(image=twentypercent)
                    enemy2healthlabel.configure(text=enemy2healthvar)
                    secenemy_secondatkcomplete=False
                    secenemy_firstatkcomplete=True
                    enemy2hasdefence=False
                else:
                    enemy2health.configure(image=fiftypercent)
                    enemy2healthlabel.configure(text=enemy2healthvar)
                    secenemy_firstatkcomplete=False
                    secenemy_secondatkcomplete=True

        except:
            print(".")

        if  secenemy_atksetup==True:
            enemy2health.configure(image=twentypercent)
            enemy2healthlabel.configure(text=enemy2healthvar)
            secenemy_firstatkcomplete=True
            secenemy_atksetup=False


        else:
            print(".")


    firstbscreen.mainloop()





def strikecard():
    global enemyhealthlabel
    global enemyhealth
    global enemyhealthvar
    global firstbscreen
    global atksetup
    global firstatkcomplete
    global secondatkcomplete
    global thirdatkcomplete
    global fourthatkcomplete
    global bg
    global enemyhasdefence
    global enemy2hasdefence
    global enemyshieldvar
    global enemyshieldbar
    global energynumber
    global doublestrength
    global thespearmanhasfallen
    global spearmanchoice
    global banditmanchoice
    global pbL
    global psLg
    global enemyintentlabel
    global selclick
    global enemy

    try:
        spearmanchoice.place_forget()
        banditmanchoice.place_forget()
        selclick=0
    except:
        print("--------------------------")
    seventypercent=PhotoImage(file="images/healthbar 70 percent dmg.png")
    ninetypercent=PhotoImage(file="images/healthbar 90 percent dmg.png")
    fiftypercent=PhotoImage(file="images/healthbar half dmg.png")
    twentypercent=PhotoImage(file="images/healthbar small dmg.png")
    onehundredpercent=PhotoImage(file="images/healthbar 100 percent dmg.png")
    reward=PhotoImage(file="images/win screen.png")
    if doublestrength==True:
        attackpower=6*2
        doublestrength=False

    else:
        attackpower=6

    if energynumber <0:
        print("cannot attack")

    else:
        pygame.mixer.init()
        pygame.mixer.music.load("Audio/cartoon punch 1.mp3")
        pygame.mixer.music.play(loops=1)
        enemyhealthvar=enemyhealthvar-attackpower

        if enemyhealthvar <= 0:
            thespearmanhasfallen=True
            Deadintent = PhotoImage(file="images/Dead mob.png")
            enemyintentlabel.config(image=Deadintent)
            enemy.place_forget()
            enemyhealthlabel.place_forget()
            enemyhealth.place_forget()
            RewardScreen()

        try:
            if enemyhasdefence==True:
                try:
                    enemyshieldvar.place_forget()
                    enemyshieldbar.place_forget()

                except:
                    print("bandit shield not spawned")
            if enemy2hasdefence==True:
                try:
                    secondenemyshieldvar.place_forget()
                    secondenemyshieldbar.place_forget()

                except:
                    print("bandit shield not spawned")


        except:
            print("enemy has no shield")



        try:
            if fourthatkcomplete==True:

                if enemyhasdefence==True:
                    enemyhealthvar = enemyhealthvar + attackpower
                    enemyhealth.configure(image=ninetypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    fourthatkcomplete = True
                    enemyhasdefence=False

                else:
                    Deadintent=PhotoImage(file="images/Dead mob.png")
                    enemyhealth.configure(image=onehundredpercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)

                    fourthatkcomplete=False
                    thespearmanhasfallen=True
                    coinvalue=psL['B1']
                    coinvalue.value=195
                    pbL.save('Excel files/Player Data.xlsx')
                    enemyintentlabel.config(image=Deadintent)
                    RewardScreen()




        except:
            print(".")



        try:
            if thirdatkcomplete==True:

                if enemyhasdefence==True:
                    enemyhealthvar = enemyhealthvar + attackpower
                    enemyhealth.configure(image=seventypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    thirdatkcomplete=True
                    fourthatkcomplete=False
                    enemyhasdefence=False
                else:
                    from TheWorld import coinslvl2clear
                    enemyhealth.configure(image=ninetypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    thirdatkcomplete=False
                    fourthatkcomplete=True
                    firstbscreen.after(100,coinslvl2clear)



        except:
            print(".")


        try:
            if secondatkcomplete==True:

                if enemyhasdefence==True:
                    enemyhealthvar = enemyhealthvar + attackpower
                    enemyhealth.configure(image=fiftypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    secondatkcomplete=True
                    thirdatkcomplete=False
                    enemyhasdefence=False
                else:
                    enemyhealth.configure(image=seventypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    secondatkcomplete=False
                    thirdatkcomplete=True

        except:
            print(".")

        try:
            if firstatkcomplete==True:

                if enemyhasdefence==True:
                    enemyhealthvar=enemyhealthvar+attackpower
                    enemyhealth.configure(image=twentypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    secondatkcomplete=False
                    firstatkcomplete=True
                    enemyhasdefence=False
                else:
                    enemyhealth.configure(image=fiftypercent)
                    enemyhealthlabel.configure(text=enemyhealthvar)
                    firstatkcomplete=False
                    secondatkcomplete=True

        except:
            print(".")

        if  atksetup==True:
            enemyhealth.configure(image=twentypercent)
            enemyhealthlabel.configure(text=enemyhealthvar)
            firstatkcomplete=True
            atksetup=False


        else:
            print(".")


    firstbscreen.mainloop()




def TheSelector():
    global firstbscreen
    global spearmanchoice
    global banditmanchoice
    global selclick
    global energynumber
    global thebandithasfallen
    global thespearmanhasfallen
    arrow=PhotoImage(file="images/target selector.png")
    selclick=selclick+1

    if selclick ==2:
        print("please do it one card at a time")

    if selclick==1:

        if thespearmanhasfallen ==True:
            print("")
        else:
            spearmanchoice=Button(firstbscreen,image=arrow,command=strikecard)
            spearmanchoice.place(relx=0.8,y=160)

        if thebandithasfallen==True:
            print("")
        else:
            banditmanchoice=Button(firstbscreen,image=arrow,command=strikecardforsecondenemy)
            banditmanchoice.place(relx=0.575,y=240)


    firstbscreen.mainloop()

def TheSelectorSPE():
    global firstbscreen
    global spearmanchoice
    global banditmanchoice
    arrow=PhotoImage(file="images/target selector.png")

    spearmanchoice=Button(firstbscreen,image=arrow,command=heavycard)
    spearmanchoice.place(relx=0.8,y=160)

    banditmanchoice=Button(firstbscreen,image=arrow,command=heavycard_secondenenmy)
    banditmanchoice.place(relx=0.575,y=240)

    firstbscreen.mainloop()

def deckcreation():
    global atkcardcounter
    global blockcardcounter
    global Drawlabel
    decktotal=atkcardcounter+blockcardcounter

    if decktotal <0:
        decktotal=3

    Drawlabel.configure(text=decktotal)




def discardpile5():
    global slot5
    global Energylabel
    global discardlabel
    global energynumber
    global discardnumber
    global slot5cardplayed

    slot5cardplayed=True
    if energynumber > 0:
        slot5.place_forget()
        energynumber=energynumber-1
        print("your energy number is", energynumber)
        discardnumber=discardnumber+1
        Energylabel.configure(text=(energynumber,"/3"))
        discardlabel.configure(text=discardnumber)
    else:
        energynumber=-5
        print("you got no energy!!!")



def discardpile4():
    global slot4
    global Energylabel
    global discardlabel
    global energynumber
    global discardnumber
    global slot4cardplayed

    slot4cardplayed=True
    if energynumber > 0:
        slot4.place_forget()
        energynumber=energynumber-1
        print("your energy number is",energynumber)
        discardnumber=discardnumber+1
        Energylabel.configure(text=(energynumber,"/3"))
        discardlabel.configure(text=discardnumber)
    else:
        energynumber = -5
        print("you got no energy !!!")



def discardpile3():
    global slot3
    global Energylabel
    global discardlabel
    global energynumber
    global discardnumber
    global slot3cardplayed

    slot3cardplayed=True
    if energynumber > 0:
        slot3.place_forget()
        energynumber=energynumber-1
        print("your energy number is", energynumber)
        discardnumber=discardnumber+1
        Energylabel.configure(text=(energynumber,"/3"))
        discardlabel.configure(text=discardnumber)
    else:
        energynumber = -5
        print("you got no energy!!!")


def discardpile2():
    global slot2
    global Energylabel
    global discardlabel
    global energynumber
    global discardnumber
    global slot2cardplayed


    slot2cardplayed=True
    if energynumber > 0:
        slot2.place_forget()
        energynumber=energynumber-1
        print("your energy number is", energynumber)
        discardnumber=discardnumber+1
        Energylabel.configure(text=(energynumber,"/3"))
        discardlabel.configure(text=discardnumber)
    else:
        energynumber = -5
        print("you got not energy!!!")



def discardpile():
    global slot1
    global Energylabel
    global discardlabel
    global energynumber
    global discardnumber
    global slot1cardplayed
    global ws2
    global wb2
    global extracost


    slot1cardplayed=True

    if energynumber >0:
        slot1.place_forget() # first card wipe process ------


        if extracost==True:
            energynumber=energynumber-2
            extracost=False

        else:
            energynumber = energynumber - 1


        discardnumber=discardnumber+1


        Energylabel.configure(text=(energynumber,"/3"))
        discardlabel.configure(text=discardnumber)
    else:
        energynumber = -5
        print("you got no Energy!!!")



def auto_drawhand_animslot5():
    global stre
    global bloEX
    global strEX
    global firstbscreen
    global atkcardcounter
    global blockcardcounter
    global blo
    global slot5
    global ws2
    global wb2
    global bigblock
    global repair
    global heavyhit
    global multihit
    global doublestr
    global more_energy
    global extracost
    global diceslot4
    global diceslot5
    global diceslot3
    global diceslot2
    global diceslot1

    if ws2['A10'].value == "repair":
        print("special dice rolled")
        if diceslot4 or diceslot3 or diceslot2 or diceslot1 == 3:
            diceslot5 = random.randint(1, 2)
        else:
            diceslot5 = random.randint(1, 3)


    elif ws2['A10'].value == "multi":
        print("special dice rolled")
        if diceslot4 or diceslot3 or diceslot2 or diceslot1 == 3:
            diceslot5 = random.randint(1, 2)
        else:
            diceslot5 = random.randint(1, 3)

    elif ws2['A10'].value == "heavy":
        print("special dice rolled")
        if diceslot4 or diceslot3 or diceslot2 or diceslot1 == 3:
            diceslot5 = random.randint(1, 2)
        else:
            diceslot5 = random.randint(1, 3)

    elif ws2['A10'].value == "30block":
        print("special dice rolled")
        if diceslot4 or diceslot3 or diceslot2 or diceslot1 == 3:
            diceslot5 = random.randint(1, 2)
        else:
            diceslot5 = random.randint(1, 3)

    elif ws2['A10'].value == "moreE":
        print("special dice rolled")
        if diceslot4 or diceslot3 or diceslot2 or diceslot1 == 3:
            diceslot5 = random.randint(1, 2)
        else:
            diceslot5 = random.randint(1, 3)

    elif ws2['A10'].value == "double":
        print("special dice rolled")
        if diceslot4 or diceslot3 or diceslot2 or diceslot1 == 3:
            diceslot5 = random.randint(1, 2)
        else:
            diceslot5 = random.randint(1, 3)

    else:
        print("normal dice")
        diceslot5 = random.randint(1, 2)



    if diceslot5 == 1:


        if wsL['A2'].value =="Reskin":
            slot5 = Button(firstbscreen, image=strEX, command=lambda: [discardpile5(), TheSelector()])
            slot5.place(relx=0.6, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

        else:
            slot5 = Button(firstbscreen, image=stre,command=lambda:[discardpile5(),TheSelector()])
            slot5.place(relx=0.6, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

    if diceslot5 == 2:
        if wsL['A2'].value =="Reskin":
            slot5 = Button(firstbscreen, image=bloEX, command=lambda: [discardpile5(), defencecard()])
            slot5.place(relx=0.6, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

        else:
            slot5 = Button(firstbscreen, image=blo,command=lambda:[discardpile5(),defencecard()])
            slot5.place(relx=0.6, y=800, anchor="center")
            blockcardcounter = blockcardcounter - 1

    if diceslot5 == 3:
        if ws2['A10'].value == "repair":
            slot5 = Button(firstbscreen, image=repair, command=lambda: [discardpile5(), repaircard()])
            slot5.place(relx=0.6, y=800, anchor="center")

        if ws2['A10'].value == "30block":
            slot5 = Button(firstbscreen, image=bigblock, command=lambda: [discardpile5(), bigblockcard()])
            slot5.place(relx=0.6, y=800, anchor="center")
            extracost = True

        if ws2['A10'].value == "multi":
            slot5 = Button(firstbscreen, image=multihit, command=lambda: [discardpile5(), multihitcard()])
            slot5.place(relx=0.6, y=800, anchor="center")

        if ws2['A10'].value == "moreE":
            slot5 = Button(firstbscreen, image=more_energy, command=lambda: [discardpile5(), moreEcard()])
            slot5.place(relx=0.6, y=800, anchor="center")

        if ws2['A10'].value == "double":
            slot5 = Button(firstbscreen, image=doublestr, command=lambda: [discardpile5(), doublestrcard()])
            slot5.place(relx=0.6, y=800, anchor="center")

        if ws2['A10'].value == "heavy":
            slot5 = Button(firstbscreen, image=heavyhit, command=lambda: [discardpile5(), TheSelectorSPE()])
            slot5.place(relx=0.6, y=800, anchor="center")
            extracost = True
    firstbscreen.mainloop()

    print("atk card in deck now are",atkcardcounter)
    print("block cards in deck now are",blockcardcounter)
    deckcreation()
    firstbscreen.mainloop()


def auto_drawhand_animslot4():
    global stre
    global blo
    global strEX
    global bloEX
    global firstbscreen
    global atkcardcounter
    global blockcardcounter
    global slot4
    global ws2
    global wb2
    global bigblock
    global repair
    global heavyhit
    global multihit
    global doublestr
    global more_energy
    global extracost
    global diceslot3
    global diceslot4
    global diceslot2
    global diceslot1

    if ws2['A10'].value == "repair":
        print("special dice rolled")
        if diceslot3 or diceslot2 or diceslot1 == 3:
            diceslot4 = random.randint(1, 2)
        else:
            diceslot4 = random.randint(1, 3)


    elif ws2['A10'].value == "multi":
        print("special dice rolled")
        if diceslot3 or diceslot2 or diceslot1 == 3:
            diceslot4 = random.randint(1, 2)
        else:
            diceslot4 = random.randint(1, 3)

    elif ws2['A10'].value == "heavy":
        print("special dice rolled")
        if diceslot3 or diceslot2 or diceslot1 == 3:
            diceslot4 = random.randint(1, 2)
        else:
            diceslot4 = random.randint(1, 3)

    elif ws2['A10'].value == "30block":
        print("special dice rolled")
        if diceslot3 or diceslot2 or diceslot1 == 3:
            diceslot4 = random.randint(1, 2)
        else:
            diceslot4 = random.randint(1, 3)

    elif ws2['A10'].value == "moreE":
        print("special dice rolled")
        if diceslot3 or diceslot2 or diceslot1 == 3:
            diceslot4 = random.randint(1, 2)
        else:
            diceslot4 = random.randint(1, 3)

    elif ws2['A10'].value == "double":
        print("special dice rolled")
        if diceslot3 or diceslot2 or diceslot1 == 3:
            diceslot4 = random.randint(1, 2)
        else:
            diceslot4 = random.randint(1, 3)

    else:
        print("normal dice")
        diceslot4 = random.randint(1, 2)


    if diceslot4 == 1:

        if wsL['A2'].value =="Reskin":
            slot4 = Button(firstbscreen, image=strEX, command=lambda: [discardpile4(), TheSelector()])
            slot4.place(relx=0.5, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

        else:
            slot4 = Button(firstbscreen, image=stre,command=lambda:[discardpile4(),TheSelector()])
            slot4.place(relx=0.5, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

    if diceslot4 == 2:

        if wsL['A2'].value =="Reskin":
            slot4 = Button(firstbscreen, image=bloEX, command=lambda: [discardpile4(), defencecard()])
            slot4.place(relx=0.5, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

        else:
            slot4 = Button(firstbscreen, image=blo,command=lambda:[discardpile4(),defencecard()])
            slot4.place(relx=0.5, y=800, anchor="center")
            blockcardcounter = blockcardcounter - 1

    if diceslot4 == 3:
        if ws2['A10'].value == "repair":
            slot4 = Button(firstbscreen, image=repair, command=lambda: [discardpile4(), repaircard()])
            slot4.place(relx=0.5, y=800, anchor="center")

        if ws2['A10'].value == "30block":
            slot4 = Button(firstbscreen, image=bigblock, command=lambda: [discardpile4(), bigblockcard()])
            slot4.place(relx=0.5, y=800, anchor="center")
            extracost = True

        if ws2['A10'].value == "multi":
            slot4 = Button(firstbscreen, image=multihit, command=lambda: [discardpile4(), multihitcard()])
            slot4.place(relx=0.5, y=800, anchor="center")

        if ws2['A10'].value == "moreE":
            slot4 = Button(firstbscreen, image=more_energy, command=lambda: [discardpile4(), moreEcard()])
            slot4.place(relx=0.5, y=800, anchor="center")

        if ws2['A10'].value == "double":
            slot4 = Button(firstbscreen, image=doublestr, command=lambda: [discardpile4(), doublestrcard()])
            slot4.place(relx=0.5, y=800, anchor="center")

        if ws2['A10'].value == "heavy":
            slot4 = Button(firstbscreen, image=heavyhit, command=lambda: [discardpile4(), TheSelectorSPE()])
            slot4.place(relx=0.5, y=800, anchor="center")
            extracost = True
    firstbscreen.mainloop()


def auto_drawhand_animslot3():
    global stre
    global blo
    global bloEX
    global strEX
    global firstbscreen
    global atkcardcounter
    global blockcardcounter
    global slot3
    global ws2
    global wb2
    global bigblock
    global repair
    global heavyhit
    global multihit
    global doublestr
    global more_energy
    global extracost
    global diceslot2
    global diceslot3
    global diceslot1
    if ws2['A10'].value == "repair":
        print("special dice rolled")
        if diceslot2 or diceslot1 == 3:
            diceslot3 = random.randint(1, 2)
        else:
            diceslot3 = random.randint(1, 3)


    elif ws2['A10'].value == "multi":
        print("special dice rolled")
        if diceslot2 or diceslot1 == 3:
            diceslot3 = random.randint(1, 2)
        else:
            diceslot3 = random.randint(1, 3)

    elif ws2['A10'].value == "heavy":
        print("special dice rolled")
        if diceslot2 or diceslot1 == 3:
            diceslot3 = random.randint(1, 2)
        else:
            diceslot3 = random.randint(1, 3)

    elif ws2['A10'].value == "30block":
        print("special dice rolled")
        if diceslot2 or diceslot1 == 3:
            diceslot3 = random.randint(1, 2)
        else:
            diceslot3 = random.randint(1, 3)

    elif ws2['A10'].value == "moreE":
        print("special dice rolled")
        if diceslot2 or diceslot1 == 3:
            diceslot3 = random.randint(1, 2)
        else:
            diceslot3 = random.randint(1, 3)

    elif ws2['A10'].value == "double":
        print("special dice rolled")
        if diceslot2 or diceslot1 == 3:
            diceslot3 = random.randint(1, 2)
        else:
            diceslot3 = random.randint(1, 3)

    else:
        print("normal dice")
        diceslot3 = random.randint(1, 2)



    if diceslot3 == 1:

        if wsL['A2'].value =="Reskin":
            slot3 = Button(firstbscreen, image=strEX, command=lambda: [discardpile3(), TheSelector()])
            slot3.place(relx=0.4, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

        else:
            slot3 = Button(firstbscreen, image=stre,command=lambda:[discardpile3(),TheSelector()])
            slot3.place(relx=0.4, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

    if diceslot3 == 2:
        if wsL['A2'].value =="Reskin":
            slot3 = Button(firstbscreen, image=bloEX, command=lambda: [discardpile3(), defencecard()])
            slot3.place(relx=0.4, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1
        else:
            slot3 = Button(firstbscreen, image=blo,command=lambda:[discardpile3(),defencecard()])
            slot3.place(relx=0.4, y=800, anchor="center")
            blockcardcounter = blockcardcounter - 1

    if diceslot3 == 3:
        if ws2['A10'].value == "repair":
            slot3 = Button(firstbscreen, image=repair, command=lambda: [discardpile3(), repaircard()])
            slot3.place(relx=0.4, y=800, anchor="center")

        if ws2['A10'].value == "30block":
            slot3 = Button(firstbscreen, image=bigblock, command=lambda: [discardpile3(), bigblockcard()])
            slot3.place(relx=0.4, y=800, anchor="center")
            extracost = True

        if ws2['A10'].value == "multi":
            slot3 = Button(firstbscreen, image=multihit, command=lambda: [discardpile3(), multihitcard()])
            slot3.place(relx=0.4, y=800, anchor="center")

        if ws2['A10'].value == "moreE":
            slot3 = Button(firstbscreen, image=more_energy, command=lambda: [discardpile3(), moreEcard()])
            slot3.place(relx=0.4, y=800, anchor="center")

        if ws2['A10'].value == "double":
            slot3 = Button(firstbscreen, image=doublestr, command=lambda: [discardpile3(), doublestrcard()])
            slot3.place(relx=0.4, y=800, anchor="center")

        if ws2['A10'].value == "heavy":
            slot3 = Button(firstbscreen, image=heavyhit, command=lambda: [discardpile3(), TheSelectorSPE()])
            slot3.place(relx=0.4, y=800, anchor="center")
            extracost = True

    firstbscreen.mainloop()

def auto_drawhand_animslot2():
    global stre
    global strEX
    global firstbscreen
    global blo
    global bloEX
    global atkcardcounter
    global blockcardcounter
    global slot2
    global ws2
    global wb2
    global bigblock
    global repair
    global heavyhit
    global multihit
    global doublestr
    global more_energy
    global extracost
    global diceslot1
    global diceslot2

    if ws2['A10'].value == "repair":
        print("special dice rolled")
        if diceslot1 == 3:
            diceslot2 = random.randint(1,2)
        else:
            diceslot2 = random.randint(1, 3)


    elif ws2['A10'].value == "multi":
        print("special dice rolled")
        if diceslot1 == 3:
            diceslot2 = random.randint(1,2)
        else:
            diceslot2 = random.randint(1, 3)

    elif ws2['A10'].value == "heavy":
        print("special dice rolled")
        if diceslot1 == 3:
            diceslot2 = random.randint(1,2)
        else:
            diceslot2 = random.randint(1, 3)

    elif ws2['A10'].value == "30block":
        print("special dice rolled")
        if diceslot1 == 3:
            diceslot2 = random.randint(1,2)
        else:
            diceslot2 = random.randint(1, 3)

    elif ws2['A10'].value == "moreE":
        print("special dice rolled")
        if diceslot1 == 3:
            diceslot2 = random.randint(1,2)
        else:
            diceslot2 = random.randint(1, 3)

    elif ws2['A10'].value == "double":
        print("special dice rolled")
        if diceslot1 == 3:
            diceslot2 = random.randint(1,2)
        else:
            diceslot2 = random.randint(1, 3)

    else:
        print("normal dice")
        diceslot2 = random.randint(1, 2)


    if diceslot2 == 1:
        if wsL['A2'].value =="Reskin":
            slot2 = Button(firstbscreen, image=strEX, command=lambda: [discardpile2(), TheSelector()])
            slot2.place(relx=0.3, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1
        else:
            slot2 = Button(firstbscreen, image=stre,command=lambda:[discardpile2(),TheSelector()])
            slot2.place(relx=0.3, y=800, anchor="center")
            atkcardcounter = atkcardcounter-1


    if diceslot2 == 2:

        if wsL['A2'].value =="Reskin":
            slot2 = Button(firstbscreen, image=bloEX, command=lambda: [discardpile2(), defencecard()])
            slot2.place(relx=0.3, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

        else:
            slot2 = Button(firstbscreen, image=blo,command=lambda:[discardpile2(),defencecard()])
            slot2.place(relx=0.3, y=800, anchor="center")
            blockcardcounter = blockcardcounter-1

    if diceslot2 == 3:
        if ws2['A10'].value == "repair":
            slot2 = Button(firstbscreen, image=repair, command=lambda: [discardpile2(), repaircard()])
            slot2.place(relx=0.3, y=800, anchor="center")

        if ws2['A10'].value == "30block":
            slot2 = Button(firstbscreen, image=bigblock, command=lambda: [discardpile2(), bigblockcard()])
            slot2.place(relx=0.3, y=800, anchor="center")
            extracost = True

        if ws2['A10'].value == "multi":
            slot2 = Button(firstbscreen, image=multihit, command=lambda: [discardpile2(), multihitcard()])
            slot2.place(relx=0.3, y=800, anchor="center")

        if ws2['A10'].value == "moreE":
            slot2 = Button(firstbscreen, image=more_energy, command=lambda: [discardpile2(), moreEcard()])
            slot2.place(relx=0.3, y=800, anchor="center")

        if ws2['A10'].value == "double":
            slot2 = Button(firstbscreen, image=doublestr, command=lambda: [discardpile2(), doublestrcard()])
            slot2.place(relx=0.3, y=800, anchor="center")

        if ws2['A10'].value == "heavy":
            slot2 = Button(firstbscreen, image=heavyhit, command=lambda: [discardpile2(), TheSelectorSPE])
            slot2.place(relx=0.3, y=800, anchor="center")
            extracost = True


    firstbscreen.mainloop()

def auto_drawhand_animslot1():
    global stre
    global firstbscreen
    global blo
    global atkcardcounter
    global blockcardcounter
    global slot1
    global wsL
    global ws2
    global strEX
    global bigblock
    global repair
    global heavyhit
    global multihit
    global doublestr
    global more_energy
    global extracost
    global diceslot1

    if ws2['A10'].value == "repair":
        print("special dice rolled")
        diceslot1=random.randint(1,3)


    elif ws2['A10'].value == "multi":
        print("special dice rolled")
        diceslot1=random.randint(1,3)

    elif ws2['A10'].value == "heavy":
        print("special dice rolled")
        diceslot1=random.randint(1,3)

    elif ws2['A10'].value == "30block":
        print("special dice rolled")
        diceslot1=random.randint(1,3)

    elif ws2['A10'].value == "moreE":
        print("special dice rolled")
        diceslot1=random.randint(1,3)

    elif ws2['A10'].value == "double":
        print("special dice rolled")
        diceslot1=random.randint(1,3)

    else:
        print("normal dice")
        diceslot1=random.randint(1,2)


    if diceslot1 ==1:

        if wsL['A2'].value =="Reskin":
            slot1 = Button(firstbscreen, image=strEX, command=lambda: [discardpile(), TheSelector()])
            slot1.place(relx=0.2, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

        else:
            slot1 = Button(firstbscreen, image=stre,command=lambda:[discardpile(),TheSelector()])
            slot1.place(relx=0.2, y=800, anchor="center")
            atkcardcounter=atkcardcounter-1




    if diceslot1 ==2:

        if wsL['A2'].value =="Reskin":
            slot1 = Button(firstbscreen, image=bloEX, command=lambda: [discardpile(), defencecard()])
            slot1.place(relx=0.2, y=800, anchor="center")
            atkcardcounter = atkcardcounter - 1

        else:
            slot1 = Button(firstbscreen, image=blo,command=lambda:[discardpile(),defencecard()])
            slot1.place(relx=0.2, y=800, anchor="center")
            blockcardcounter=blockcardcounter-1


    if diceslot1 ==3:
        if ws2['A10'].value =="repair":
            slot1 = Button(firstbscreen, image=repair,command=lambda:[discardpile(),repaircard()])
            slot1.place(relx=0.2, y=800, anchor="center")

        if ws2['A10'].value =="30block":
            slot1 = Button(firstbscreen, image=bigblock,command=lambda:[discardpile(),bigblockcard()])
            slot1.place(relx=0.2, y=800, anchor="center")
            extracost=True

        if ws2['A10'].value =="multi":
            slot1 = Button(firstbscreen, image=multihit,command=lambda:[discardpile(),multihitcard()])
            slot1.place(relx=0.2, y=800, anchor="center")

        if ws2['A10'].value =="moreE":
            slot1 = Button(firstbscreen, image=more_energy,command=lambda:[discardpile(),moreEcard()])
            slot1.place(relx=0.2, y=800, anchor="center")

        if ws2['A10'].value =="double":
            slot1 = Button(firstbscreen, image=doublestr,command=lambda:[discardpile(),doublestrcard()])
            slot1.place(relx=0.2, y=800, anchor="center")

        if ws2['A10'].value =="heavy":
            slot1 = Button(firstbscreen, image=heavyhit,command=lambda:[discardpile(),TheSelectorSPE()])
            slot1.place(relx=0.2, y=800, anchor="center")
            extracost=True


    firstbscreen.mainloop()

def autodrawhand():
    global ws2
    global firstbscreen
    global anim2
    global stre
    global blo
    global blockcardcounter
    global atkcardcounter
    global strEX
    global bloEX
    global bigblock
    global repair
    global heavyhit
    global multihit
    global doublestr
    global more_energy
    blockcardcounter=0
    atkcardcounter=0
    stre= PhotoImage(file="images/atk card.png")
    strEX = PhotoImage(file="images/atk card reskin.png")
    blo = PhotoImage(file="images/block card.png")
    bloEX = PhotoImage(file="images/atk card reskin.png")
    bigblock=PhotoImage(file="images/30 block card.png")
    repair=PhotoImage(file="images/Self repair card.png")
    heavyhit=PhotoImage(file="images/Heavy hit card reskin.png")
    multihit=PhotoImage(file="images/multi slash card.png")
    doublestr=PhotoImage(file="images/Double Strength card.png")
    more_energy=PhotoImage(file="images/Double Energy.png")

    if ws2['A2'].value == "Strike":
        atkcardcounter=atkcardcounter+1

    if ws2['A2'].value == "Block":
        blockcardcounter = blockcardcounter+1

    if ws2['A3'].value == "Strike":
        atkcardcounter = atkcardcounter + 1


    if ws2['A3'].value == "Block":
        blockcardcounter = blockcardcounter + 1


    if ws2['A4'].value == "Strike":
        atkcardcounter = atkcardcounter + 1


    if ws2['A4'].value == "Block":
        blockcardcounter = blockcardcounter + 1


    if ws2['A5'].value == "Strike":
        atkcardcounter = atkcardcounter + 1


    if ws2['A5'].value == "Block":
        blockcardcounter = blockcardcounter + 1


    if ws2['A6'].value == "Strike":
        atkcardcounter = atkcardcounter + 1

    if ws2['A6'].value == "Block":
        blockcardcounter = blockcardcounter + 1

    if ws2['A7'].value == "Strike":
        atkcardcounter = atkcardcounter + 1

    if ws2['A7'].value == "Block":
        blockcardcounter = blockcardcounter + 1


    if ws2['A8'].value == "Strike":
        atkcardcounter = atkcardcounter + 1

    if ws2['A8'].value == "Block":
        blockcardcounter = blockcardcounter + 1

    if ws2['A9'].value == "Strike":
        atkcardcounter = atkcardcounter + 1

    if ws2['A9'].value == "Block":
        blockcardcounter = blockcardcounter + 1

    print("atk cards in Deck",atkcardcounter)
    print("Block cards in Deck",blockcardcounter)


    firstbscreen.after(900,auto_drawhand_animslot1)
    firstbscreen.after(1400, auto_drawhand_animslot2)
    firstbscreen.after(1900, auto_drawhand_animslot3)
    firstbscreen.after(2400, auto_drawhand_animslot4)
    firstbscreen.after(2900, auto_drawhand_animslot5)



def Endturnreset():
    global discardnumber
    global energynumber
    global Energylabel
    global discardlabel
    global slot1
    global slot2
    global slot3
    global slot4
    global slot5
    global slot1cardplayed
    global slot2cardplayed
    global slot3cardplayed
    global slot4cardplayed
    global slot5cardplayed
    global Shieldvar
    global shieldlabel
    global shieldnum
    global firstbscreen
    global enemyhasdefence
    global enemyshieldbar
    global enemyshieldvar
    global enemy2hasdefence
    global secondenemyshieldvar
    global secondenemyshieldbar


    energynumber=3
    Energylabel.configure(text=(energynumber,"/3"))
# if shield have not been defined then the except will be played -----

    try:
        secondenemyshieldvar.place_forget()
        secondenemyshieldbar.place_forget()
        enemy2hasdefence=False
    except:
        print("enemy 2 shields were not reset")


    try:
        enemyshieldvar.place_forget()
        enemyshieldbar.place_forget()
        enemyhasdefence=False
    except:
        print("enemy 1 shields were not reset")


    try:
        enemyshieldvar.place_forget()
        enemyshieldbar.place_forget()
        enemyhasdefence=False
    except:
        print("enemy 1 shields were not reset")

    try:
        Shieldvar=0
        shieldnum.place_forget()
        shieldlabel.place_forget()
        enemyhasdefence=False
    except:
        print("shields were not reset")

    try:
        if slot1cardplayed==True:
            slot1cardplayed=False
    except:
        print("slot 1 cleared")
        discardnumber = discardnumber + 1
        firstbscreen.after(100,slot1.destroy)

    else:
        slot1.place_forget()
    try:
        if slot2cardplayed==True:
            slot2cardplayed=False

    except:
        print("slot 2 cleared")
        firstbscreen.after(100, slot2.destroy)
        discardnumber=discardnumber+1

    else:
        slot2.place_forget()

    try:
        if slot3cardplayed==True:
            slot3cardplayed=False

    except:
        print("slot 3 cleared")
        firstbscreen.after(100, slot3.destroy)
        discardnumber=discardnumber+1

    else:
        slot3.place_forget()

    try:
        if slot4cardplayed==True:
            slot4cardplayed=False
    except:
        print("slot 4 cleared")
        firstbscreen.after(100, slot4.destroy)
        discardnumber=discardnumber+1

    else:
        slot4.place_forget()
    try:
        if slot5cardplayed==True:
            slot5cardplayed=False
    except:
        print("slot 5 cleared")
        firstbscreen.after(100, slot5.destroy)
        discardnumber=discardnumber+1

    else:
        slot5.place_forget()




    discardlabel.configure(text=discardnumber)
    firstbscreen.after(900,auto_drawhand_animslot1)
    firstbscreen.after(1400, auto_drawhand_animslot2)
    firstbscreen.after(1900, auto_drawhand_animslot3)
    firstbscreen.after(2400, auto_drawhand_animslot4)
    firstbscreen.after(2900, auto_drawhand_animslot5)




class ImageLabel(Label):
    """a label that displays images, and plays them if they are gifs"""
    def load(self, im):
        if isinstance(im, str):
            im = Image.open(im)
        self.loc = 0
        self.frames = []

        try:
            for i in count(1):
                self.frames.append(ImageTk.PhotoImage(im.copy()))
                im.seek(i)
        except EOFError:
            pass

        try:
            self.delay = im.info['duration']
        except:
            self.delay = 100

        if len(self.frames) == 1:
            self.config(image=self.frames[0])
        else:
            self.next_frame()

    def unload(self):
        self.config(image="")
        self.frames = None

    def next_frame(self):
        if self.frames:
            self.loc += 1
            self.loc %= len(self.frames)
            self.config(image=self.frames[self.loc])
            self.after(self.delay, self.next_frame)

def background_Track():
    pygame.mixer.music.stop()

def LanguageCheck():
    global wsL
    global wbL
    global Endturn
    if wsL['A1'].value =="Bengali":
        Endturn.config(text="এন্ড টার্ন")

    if wsL['A1'].value =="Japanese":
        Endturn.config(text="エンドターン")

    if wsL['A1'].value =="English":
        Endturn.config(text="End Turn")


def gamesetup():
    global firstbscreen
    global mainguy
    global Drawlabel
    global Energylabel
    global discardlabel
    global energynumber
    global discardnumber
    global enemyhealth
    global enemyhealthlabel
    global enemyhealthvar
    global enemy2health
    global enemy2healthlabel
    global enemy2healthvar
    global atksetup
    global secenemy_atksetup
    global health
    global healthlabel
    global healthvar
    global enemyatksetup
    global Shieldvar
    global cardclicked
    global shieldactivated
    global enemy
    global bg
    global enemyintentlabel
    global secondenemyintent
    global enemyhasdefence
    global Endturn
    global doublestrength
    global extracost
    global thespearmanhasfallen
    global thebandithasfallen
    global selclick
    global enemy
    global enemy2
    global psL
    global pbL
    thespearmanhasfallen=False
    thebandithasfallen=False
    doublestrength=False
    extracost=False
    bg = PhotoImage(file="images/background.png")
    dk = PhotoImage(file="images/deck.png")
    ds = PhotoImage(file="images/discard.png")
    en = PhotoImage(file="images/enemy.png")
    ener = PhotoImage(file="images/energy.png")
    mhb = PhotoImage(file="images/healthbar.png")
    ehb = PhotoImage(file="images/healthbar.png")
    qint = PhotoImage(file="images/unkown intent.PNG")
    coincolourbg = '#%02x%02x%02x' % (61,73,79)




    energynumber=3
    discardnumber=0
    enemyhealthvar=30
    enemy2healthvar=30
    healthvar=psL['B2'].value
    Shieldvar=0
    cardclicked=0
    selclick=0
    atksetup=True
    enemyatksetup=True
    secenemy_atksetup=True
    shieldactivated=False
    enemyhasdefence=False

    # Show image using label
    label1 = Label(firstbscreen, image=bg)
    label1.place(x=0, y=0)

    mainguy=ImageLabel(firstbscreen)
    mainguy.place(relx=0.2, y=500, anchor="center")
    if wcS['A1'].value == "Defect":
        mainguy.load('images/The Defect.gif')
    if wcS['A1'].value == "Silent":
        mainguy.load("images/thief.gif")
    if wcS['A1'].value == "Ironclad":
        mainguy.load("images/Ironclad.gif")


    coinbag = PhotoImage(file="images/coinbag.png")
    coinbaglabel = Label(firstbscreen, image=coinbag)
    coinbaglabel.place(relx=0, y=0)
    coins = Label(firstbscreen, text="075", font=("Times New Roman", 30,), bg=coincolourbg, fg="yellow")
    coins.place(relx=0.05, y=2)


    enemy = ImageLabel(firstbscreen)
    enemy.place(relx=0.8, y=400, anchor="center")
    enemy.load("images/spearman.gif")

    enemy2 = ImageLabel(firstbscreen)
    enemy2.place(relx=0.6, y=480, anchor="center")
    enemy2.load("images/spearman.gif")


    thedeck=Label(firstbscreen, image=dk)
    thedeck.place(relx=0.1, y=800, anchor="center")

    thediscard=Label(firstbscreen, image=ds)
    thediscard.place(relx=0.8, y=800, anchor="center")

    discardlabel = Label(firstbscreen, text=discardnumber, fg="blue",font=("Arial", 25))
    discardlabel.place(relx=0.8, y=800, anchor="center")

    Endturn = Button(firstbscreen,text="End Turn",bg="Dark Blue",fg="Grey",command=lambda:[Endturnreset(),enemydecision()])
    Endturn.place(relx=0.8, y=700, anchor="center")

    Drawlabel = Label(firstbscreen, text="X", fg="blue",font=("Arial", 25))
    Drawlabel.place(relx=0.1, y=800, anchor="center")

    Energy = Label(firstbscreen,image=ener)
    Energy.place(relx=0.07,y=650)

    Energylabel =Label(firstbscreen,text=(energynumber,"/3"),fg="green",bg="cyan",font=("Arial", 35))
    Energylabel.place(relx=0.075,y=670)

    health = Label(firstbscreen,image=mhb)
    health.place(relx=0.135,y=600)

    healthlabel = Label(firstbscreen,text=(healthvar,"/",healthvar),font=("Arial", 8))
    healthlabel.place(relx=0.126,y=600)

    enemyhealth = Label(firstbscreen,image=ehb)
    enemyhealth.place(relx=0.73,y=500)
    enemyhealthlabel = Label(firstbscreen, text=(enemyhealthvar,"/",enemyhealthvar), font=("Arial", 8))
    enemyhealthlabel.place(relx=0.72, y=500)

    enemy2health = Label(firstbscreen,image=ehb)
    enemy2health.place(relx=0.53,y=550)
    enemy2healthlabel = Label(firstbscreen, text=(enemy2healthvar,"/",enemy2healthvar), font=("Arial", 8))
    enemy2healthlabel.place(relx=0.53, y=550)

    secondenemyintent = Label(firstbscreen, image=qint)
    secondenemyintent.place(relx=0.575,y=340)


    enemyintentlabel = Label(firstbscreen, image=qint)
    enemyintentlabel.place(relx=0.805,y=240)
    background_Track()
    LanguageCheck()
    autodrawhand()
    firstbscreen.mainloop()



def chestbattlestart():
    global firstbscreen
    global wb2
    global ws2
    global wbL
    global wsL
    global wcS
    global wcL
    global pbL
    global psL
    wbL = load_workbook("Excel files/Settings.xlsx")
    wsL=wbL.active

    wb2 = load_workbook("Excel files/Active Deck.xlsx")
    ws2=wb2.active

    wcL = load_workbook("Excel files/Character Selection.xlsx")
    wcS = wcL.active


    pbL = load_workbook("Excel files/Player Data.xlsx")
    psL=pbL.active


    firstbscreen=Toplevel()



    firstbscreen.geometry("1500x875")
    gamesetup()
